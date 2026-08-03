#!/usr/bin/env python
"""Build seeds/reference/budget.csv from the client's Budget and LYTD workbook.

Source: `Finance Templates/<Month> Budget and LYTD Comparison*.xlsx`, one sheet per
month (Jan, Feb, Mar, April, May, June). Each sheet is a block of three columns per
entity, with the entity name merged across them on row 1:

        |        ZAAC          |        ZARIB         | ...
        | April MTD | April MTD | April LYMTD | ...
        | Actual    | Budget    | Actual      | ...
    Actuarial Fees   27,639,193  25,000,000        0    ...

Column A carries the line label. Row 2 distinguishes MTD from LYMTD (last year, same
month) and row 3 the measure, so the third column of each block is the prior-year
actual — not part of the budget.

MTD, not YTD
------------
The budget columns are MONTH figures. Every mart in this project is cumulative
year-to-date, so a period's budget is the sum of the month budgets up to and
including it. Verified against the previously-loaded 2026-04 seed: ZAMRE, ZHL, MENA,
personnel_costs, it_costs, premises and depreciation all reproduce to the cent under
this rule.

Mapping to the Group P&L taxonomy (`report_line.csv`)
----------------------------------------------------
Income is per entity, expenses are by nature for the Kenyan entities and a single
total for the rest — mirroring how `report_line_map.csv` maps actuals:

    zaac_revenue    <- ZAAC  'Total Income'          personnel_costs .. finance_costs
    cp_revenue      <- C&P   'Total Income'              <- ZAAC + ZARIB + C&P + ZHL,
    zarib_revenue   <- ZARIB 'Total Income'                 line by line
    zhl_interest    <- ZHL   'Total Income'
    zamre_revenue   <- ZAMRE 'Total Income'          zamre_expense   <- ZAMRE 'Total Expenses'
    mena_revenue    <- MENA  'Total Income'          mena_expense    <- MENA  'Total Expenses'
    zarinet_revenue <- MALAWI + RWANDA + NIGERIA     zarinet_expense <- the same five,
                       + DRC + ZATL 'Total Income'                      'Total Expenses'

Taking each entity's own 'Total Income' rather than summing its income sub-lines
means a new income line in a later pack cannot be silently dropped.

Label drift
-----------
Jan and Feb use different wording for three lines: 'Travel' (later 'Travelling'),
'Management Fees' ('Management Expense') and 'Pension Administration' ('Pension Admin
Fee'). Normalised below. Ignoring this cost KES 3.9m on travelling alone.

'Management Expense' has no code in `report_line.csv`, which is correct — only the
African entities budget for it, and they roll up as a single total. The self-test
below fails if a Kenyan entity ever starts using it, because it would then have
nowhere to go.

Usage
-----
    python Scripts/extract_budget.py               # dry run, prints the seed + checks
    python Scripts/extract_budget.py --write       # write seeds/reference/budget.csv

The workbook's LYMTD ('last year, same month') column is deliberately NOT loaded.
Prior year should be derived from our own actuals once a full year exists, not
imported from the client's comparative — see the fiscal-year note in README.md.
"""
import argparse
import collections
import csv
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT = os.path.dirname(HERE)
WORKBOOK = os.path.join(PROJECT, os.pardir, 'Finance Templates',
                        'June Budget and LYTD Comparison - Revised (1).xlsx')
SEED = os.path.join(PROJECT, 'seeds', 'reference', 'budget.csv')
SEED_SUB = os.path.join(PROJECT, 'seeds', 'reference', 'budget_subsidiary.csv')

MONTHS = [('Jan', '2026-01'), ('Feb', '2026-02'), ('Mar', '2026-03'),
          ('April', '2026-04'), ('May', '2026-05'), ('June', '2026-06')]

# Jan/Feb wording -> the wording used from March onwards
LABEL_ALIASES = {
    'travel': 'Travelling',
    'management fees': 'Management Expense',
    'pension administration': 'Pension Admin Fee',
}

KENYA   = ['ZAAC', 'ZARIB', 'C&P', 'ZHL']
ZARINET = ['MALAWI', 'RWANDA', 'NIGERIA', 'DRC', 'ZATL']

# income: report_line_code -> the entities whose 'Total Income' it is
INCOME = {
    'zaac_revenue':    ['ZAAC'],
    'cp_revenue':      ['C&P'],
    'zarib_revenue':   ['ZARIB'],
    'zhl_interest':    ['ZHL'],
    'zamre_revenue':   ['ZAMRE'],
    'zarinet_revenue': ZARINET,
    'mena_revenue':    ['MENA'],
}

# expenses taken as a single 'Total Expenses' rather than by nature
EXPENSE_TOTALS = {
    'zamre_expense':   ['ZAMRE'],
    'zarinet_expense': ZARINET,
    'mena_expense':    ['MENA'],
}

# by-nature expense lines for the Kenyan entities: report_line_code -> sheet label
BY_NATURE = [
    ('personnel_costs',     'Personnel Costs'),
    ('travelling',          'Travelling'),
    ('entertainment',       'Entertainment'),
    ('it_costs',            'IT Costs'),
    ('communications',      'Communications'),
    ('printing_stationery', 'Printing & Stationery'),
    ('premises',            'Premises'),
    ('advertising_pr',      'Advertising & PR'),
    ('insurance',           'Insurance'),
    ('professional_fees',   'Professional Fees'),
    ('depreciation',        'Depreciation'),
    ('motor_vehicle',       'Motor Vehicle Expenses'),
    ('general_expenses',    'General Expenses'),
    ('finance_costs',       'Finance cost'),
]

# no budget is provided for these; emitted at zero so the variance column reads 0
# rather than NULL, matching how the earlier April-only seed behaved
ZERO_LINES = ['bad_debt_provision']

SUBTOTALS = {'income', 'expenses', 'total income', 'total expenses', 'profit before tax'}

# -----------------------------------------------------------------------------
# Subsidiary grain
#
# The Group P&L taxonomy above pools the four Kenyan entities into by-nature
# expense lines and the five African entities into a single Zarinet total, so
# rpt_group_pl cannot show ZAAC's personnel budget against ZAAC's personnel
# actual. The workbook itself is at entity x line grain, and all 22 of its line
# labels map 1:1 onto our SCI statement lines, so nothing was lost in that
# aggregation and a subsidiary budget needs no allocation or judgement.
#
# Only `taxation` has no budget line, which is correct — the workbook stops at
# profit before tax.
#
# The workbook is in KES: ZAAC and ZARIB, both KES-functional, reproduce our
# marts' income exactly. The foreign entities land within 0.2% (ZATL 0.8%),
# which is rate rounding between their budget rates and ours — noted rather than
# corrected, since the budget is a plan and not a translated ledger.
# -----------------------------------------------------------------------------
STATEMENT_LINE = {
    'Actuarial Fees':                               'actuarial_fees',
    'Pension Admin Fee':                            'pension_admin_fee',
    'Umbrella Fund':                                'umbrella_fund',
    'Insurance Commissions':                        'insurance_commissions',
    'Special Project-SAAS (Software as a Service)':  'saas_income',
    'Retail':                                       'retail_income',
    'Other Income':                                 'other_income',
    'Personnel Costs':                              'personnel_costs',
    'Travelling':                                   'travelling',
    'Entertainment':                                'entertainment',
    'IT Costs':                                     'it_costs',
    'Communications':                               'communications',
    'Printing & Stationery':                        'printing_stationery',
    'Premises':                                     'premises',
    'Advertising & PR':                             'advertising_pr',
    'Insurance':                                    'insurance',
    'Professional Fees':                            'professional_fees',
    'Depreciation':                                 'depreciation',
    'Motor Vehicle Expenses':                       'motor_vehicle_expenses',
    'General Expenses':                             'general_expenses',
    'Finance cost':                                 'finance_costs',
    'Management Expense':                           'management_expense',
}

# Sign: the workbook presents income positive and expenses negative. Our marts
# use the presentation basis where sign_multiplier has already flipped both to
# positive. The budget has to sit on the same basis as the actual in the same
# row or the two columns are not comparable, so expenses are negated.
INCOME_LINES = {'actuarial_fees', 'pension_admin_fee', 'umbrella_fund',
                'insurance_commissions', 'saas_income', 'retail_income', 'other_income'}

# workbook entity name -> entity.csv entity_code. Identical today; kept explicit
# so a rename in the workbook fails loudly instead of dropping an entity.
ENTITY_CODE = {e: e for e in
               ['ZAAC', 'ZARIB', 'ZAMRE', 'ZHL', 'ZATL', 'C&P', 'MENA',
                'MALAWI', 'NIGERIA', 'RWANDA', 'DRC']}


def num(v):
    """Sheet values arrive as both floats and strings ('25000000', '0', '(1,234)')."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(',', '').replace('(', '-').replace(')', '').strip()
    if s in ('', '-', 'None'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def read_workbook(path):
    """-> {(period, entity, label, measure): month_amount}

    measure is 'budget' | 'actual' | 'lymtd'. Entity columns are located from row 1
    every time rather than assumed, because the order changes between packs — Jan
    has NIGERIA before RWANDA and every later month has them the other way round.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    out = collections.defaultdict(float)
    seen_labels = collections.defaultdict(set)
    for sheet, period in MONTHS:
        if sheet not in wb.sheetnames:
            print(f'  {period}: sheet {sheet!r} not found, skipped')
            continue
        ws = wb[sheet]
        blocks = [(c, str(ws.cell(1, c).value).strip())
                  for c in range(1, ws.max_column + 1) if ws.cell(1, c).value]
        for col, entity in blocks:
            if entity.upper() == 'TOTAL':
                continue                       # the client's own cross-check column
            measures = {}
            for off in range(3):
                c = col + off
                head = str(ws.cell(3, c).value or '').strip().lower()
                span = str(ws.cell(2, c).value or '').strip().lower()
                if head == 'budget':
                    measures['budget'] = c
                elif head == 'actual':
                    measures['lymtd' if ('lymtd' in span or 'lytd' in span)
                             else 'actual'] = c
            for r in range(4, ws.max_row + 1):
                raw = ws.cell(r, 1).value
                if not raw:
                    continue
                label = str(raw).strip()
                label = LABEL_ALIASES.get(label.lower(), label)
                seen_labels[label].add(period)
                for measure, c in measures.items():
                    v = num(ws.cell(r, c).value)
                    if v is not None:
                        out[(period, entity, label, measure)] += v
    return out, seen_labels


def build(data, measure):
    """MTD -> cumulative YTD, mapped onto report_line_code."""
    def month(period, entity, label):
        return data.get((period, entity, label, measure), 0.0)

    def ytd(upto, entities, label):
        total = 0.0
        for _, period in MONTHS:
            if period > upto:
                break
            total += sum(month(period, e, label) for e in entities)
        return total

    rows = []
    for _, period in MONTHS:
        for code, entities in INCOME.items():
            rows.append((code, period, ytd(period, entities, 'Total Income')))
        for code, label in BY_NATURE:
            rows.append((code, period, ytd(period, KENYA, label)))
        for code in ZERO_LINES:
            rows.append((code, period, 0.0))
        for code, entities in EXPENSE_TOTALS.items():
            rows.append((code, period, ytd(period, entities, 'Total Expenses')))
    return rows


def build_subsidiary(data):
    """-> [(entity_code, statement_line_code, period, budget_kes)]

    No aggregation at all — one row per entity per line per period, cumulated MTD
    to YTD and sign-aligned to the marts. Nil rows are dropped so the seed does
    not carry 11 x 22 x 6 mostly-empty rows.

    The workbook's third column per entity is LYMTD ('last year, same month')
    actual, and it is deliberately NOT loaded. It would only be a stopgap: once a
    full year of our own actuals exists, prior year should be derived from
    fct_trial_balance at period minus twelve months, so the comparative equals the
    actual we published a year earlier rather than coming from the client's
    figures. Loading it now would also have to be unwound then. See the
    fiscal-year note in README.md and Project_Handover.md pt.16.
    """
    rows = []
    for _, period in MONTHS:
        for entity, code in sorted(ENTITY_CODE.items()):
            for label, line in sorted(STATEMENT_LINE.items()):
                sign = 1.0 if line in INCOME_LINES else -1.0
                bud = 0.0
                for _, p in MONTHS:
                    if p > period:
                        break
                    bud += data.get((p, entity, label, 'budget'), 0.0)
                if abs(bud) < 0.005:
                    continue
                rows.append((code, line, period, round(bud * sign, 2)))
    return rows


def check_subsidiary(sub, group):
    """The subsidiary seed must roll up to the group seed — same numbers, one
    aggregation apart. Anything else means the two disagree about what the budget
    is, which is worse than not having the subsidiary grain at all."""
    fails = []
    g = {(code, period): amount for code, period, amount in group}
    # expenses were negated for the subsidiary seed; undo that to compare with the
    # group seed, which keeps the workbook's income-positive / expense-negative basis
    by = collections.defaultdict(float)
    for code, line, period, bud in sub:
        by[(period, code, line in INCOME_LINES)] += bud * (1.0 if line in INCOME_LINES else -1.0)

    for _, period in MONTHS:
        for gcode, entities in list(INCOME.items()):
            want = g.get((gcode, period), 0.0)
            got = sum(by.get((period, e, True), 0.0) for e in entities)
            if abs(got - want) > 1.0:
                fails.append(f'{period} {gcode}: subsidiary {got:,.2f} != group {want:,.2f}')
        for gcode, entities in list(EXPENSE_TOTALS.items()):
            want = g.get((gcode, period), 0.0)
            got = sum(by.get((period, e, False), 0.0) for e in entities)
            if abs(got - want) > 1.0:
                fails.append(f'{period} {gcode}: subsidiary {got:,.2f} != group {want:,.2f}')
        # the Kenyan by-nature lines, in aggregate
        want = sum(g.get((c, period), 0.0) for c, _ in BY_NATURE)
        got = sum(by.get((period, e, False), 0.0) for e in KENYA)
        if abs(got - want) > 1.0:
            fails.append(f'{period} Kenya by-nature: subsidiary {got:,.2f} != group {want:,.2f}')
    return fails


def checks(data, rows, seen_labels, measure):
    """Returns a list of failures; empty means every self-test passed."""
    fails = []

    def month(period, entity, label):
        return data.get((period, entity, label, measure), 0.0)

    # 1. the by-nature lines must account for the whole of Kenya's Total Expenses,
    #    or an expense category is being dropped on the floor
    for _, period in MONTHS:
        by_nature = sum(month(period, e, lab) for e in KENYA for _, lab in BY_NATURE)
        stated = sum(month(period, e, 'Total Expenses') for e in KENYA)
        if abs(by_nature - stated) > 1.0:
            fails.append(f'{period}: Kenya by-nature {by_nature:,.2f} != '
                         f'Total Expenses {stated:,.2f} (diff {by_nature - stated:,.2f})')

    # 2. 'Management Expense' has no report_line_code. Fine while only the African
    #    entities use it (they roll up as a total); a problem the moment Kenya does.
    for _, period in MONTHS:
        stray = sum(month(period, e, 'Management Expense') for e in KENYA)
        if abs(stray) > 1.0:
            fails.append(f'{period}: a Kenyan entity now budgets Management Expense '
                         f'({stray:,.2f}) and it has no report_line_code')

    # 3. every non-subtotal label must be consumed by the mapping
    mapped = {lab for _, lab in BY_NATURE} | {'Total Income', 'Total Expenses',
                                              'Management Expense'}
    for label in seen_labels:
        if label.lower() in SUBTOTALS or label in mapped:
            continue
        # income sub-lines are covered by taking 'Total Income'; only flag if a label
        # sits outside both the income block and the mapped expense lines
        total = sum(month(p, e, label) for _, p in MONTHS
                    for e in KENYA + ZARINET + ['ZAMRE', 'MENA'])
        if abs(total) > 1.0 and label not in (
                'Actuarial Fees', 'Pension Admin Fee', 'Umbrella Fund',
                'Insurance Commissions', 'Special Project-SAAS (Software as a Service)',
                'Retail', 'Other Income'):
            fails.append(f'unmapped label {label!r} carries {total:,.2f}')

    # 4. YTD must be monotonic in the sense that each period's figure equals the
    #    prior period plus that month — catches a cumulation bug
    by_code = collections.defaultdict(dict)
    for code, period, amount in rows:
        by_code[code][period] = amount
    for code, series in by_code.items():
        if code in ZERO_LINES:
            continue
        periods = [p for _, p in MONTHS if p in series]
        for prev, cur in zip(periods, periods[1:]):
            entities = (INCOME.get(code) or EXPENSE_TOTALS.get(code) or KENYA)
            label = ('Total Income' if code in INCOME else
                     'Total Expenses' if code in EXPENSE_TOTALS else
                     dict((c, l) for c, l in BY_NATURE)[code])
            delta = sum(month(cur, e, label) for e in entities)
            if abs((series[prev] + delta) - series[cur]) > 1.0:
                fails.append(f'{code} {cur}: cumulation broken')
    return fails


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--write', action='store_true')
    ap.add_argument('--workbook', default=WORKBOOK)
    ap.add_argument('--seed', default=SEED)
    ap.add_argument('--seed-subsidiary', default=SEED_SUB)
    a = ap.parse_args()

    if not os.path.exists(a.workbook):
        sys.exit(f'workbook not found: {a.workbook}')
    print(f'reading {os.path.basename(a.workbook)}')
    data, seen = read_workbook(a.workbook)

    rows = build(data, 'budget')
    fails = checks(data, rows, seen, 'budget')

    sub = build_subsidiary(data)
    fails += check_subsidiary(sub, rows)

    print(f'\n{len(rows)} rows, {len({r[0] for r in rows})} report lines x '
          f'{len({r[1] for r in rows})} periods\n')
    print(f"{'report_line_code':22}" + ''.join(f'{p[-2:]:>18}' for _, p in MONTHS))
    by_code = collections.defaultdict(dict)
    for code, period, amount in rows:
        by_code[code][period] = amount
    for code in [c for c in INCOME] + [c for c, _ in BY_NATURE] + ZERO_LINES + list(EXPENSE_TOTALS):
        series = by_code.get(code, {})
        print(f'  {code:20}' + ''.join(f'{series.get(p, 0.0):>18,.0f}' for _, p in MONTHS))

    print(f'\nsubsidiary grain: {len(sub)} rows, '
          f'{len({r[0] for r in sub})} entities x {len({r[1] for r in sub})} statement lines '
          f'x {len({r[2] for r in sub})} periods')

    print('\nself-tests:')
    if fails:
        for f in fails:
            print(f'  FAIL  {f}')
    else:
        print('  all passed — Kenya by-nature ties to Total Expenses in every period, '
              'every label is consumed, cumulation is consistent, and the subsidiary '
              'seed rolls up to the group seed exactly')

    if not a.write:
        print('\ndry run — pass --write to update the seeds')
        return
    if fails:
        sys.exit('\nrefusing to write: self-tests failed')

    with open(a.seed, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['report_line_code', 'period', 'amount_budget_kes'])
        for code, period, amount in rows:
            w.writerow([code, period, round(amount, 2)])
    print(f'\nwrote {a.seed} ({len(rows)} rows)')

    with open(a.seed_subsidiary, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['company_name', 'statement_line_code', 'period', 'amount_budget_kes'])
        for r in sub:
            w.writerow(r)
    print(f'wrote {a.seed_subsidiary} ({len(sub)} rows)')
    print('\nnow run: dbt seed --select budget budget_subsidiary --full-refresh && dbt build')


if __name__ == '__main__':
    main()
