#!/usr/bin/env python3
"""
reseed_from_packs.py — rebuild the bronze gl_entry seeds from scratch out of the
client's final monthly packs (Finance Templates/2026 TBs/Consolidated Accounts).

Why this exists (and why load_tb.py cannot do it)
-------------------------------------------------
load_tb.py adds ONE new month on top of seeds that already hold every earlier
month. It computes `movement = new YTD - prior cumulative in the seed`, treating
every row that is not in the target period as "prior". Run it for January while
February-June rows are still in the seed and January's movement comes out as
`Jan YTD - (Feb..Jun movements)`. The final period still ties, but every
intermediate period is wrong. A full rebuild therefore has to start from empty
seeds and load the months in order — which is what this script does.

Emptying the seeds loses two things that have to be handled deliberately:

  1. The description -> account-code "codebook". Several tabs carry rows with a
     BLANK A/C No (ZARIB's payroll/check-off block, ZAAC 'Tax expense', Malawi
     'FIRST CAPITAL CURRENT ACCOUNT'), and two tabs are description-only (MENA,
     and Nigeria in the final packs). load_tb.py resolves those by matching the
     description against codes already in the seed — which an emptied seed
     cannot do. This script snapshots the codebook BEFORE truncating.

  2. Codes the client reuses for two different accounts. ZARIB reuses 7-10 codes
     every month (e.g. 4000/000 is both 'VALUE ADDS' and 'Leasehold-
     Depreciation'); ZATL, C&P, ZAAC and DRC each reuse one. Aggregating by code
     alone silently merges them. Splits are declared in code_overrides.csv.

Resolution order for every source row
-------------------------------------
  1. code_overrides.csv          — declared splits and pinned blank-code codes
  2. code_bridge_nigeria.csv     — Nigeria description -> BC code (Nigeria only)
  3. source code, if that code carries only one description in the packs
  4. duplicate code             -> codebook, else `<code>.2`, `.3` (flagged)
  5. blank code                 -> codebook, else `<ENT>-<SLUG>` (flagged)

MENA is keyed on the raw Description string, because stg_mena_descriptive_tb
derives its account code as 'MENA-' || md5(Description) — so the description
text is the key and must be passed through byte-for-byte.

Usage
-----
    python Scripts/reseed_from_packs.py                # dry run, writes a report
    python Scripts/reseed_from_packs.py --write        # rebuild the seeds
    python Scripts/reseed_from_packs.py --write --report-dir ../Outputs

After --write:  dbt seed --full-refresh && dbt build
Then reconcile — see Internal/Phase1_Data_Checks_and_Finance_ICT_Issues.md.
"""
import argparse
import collections
import csv
import datetime
import hashlib
import json
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT = os.path.normpath(os.path.join(HERE, '..'))
SEEDS = os.path.join(PROJECT, 'seeds')
BRONZE = os.path.join(SEEDS, 'bronze')
REF = os.path.join(SEEDS, 'reference')
PACKS = os.path.normpath(os.path.join(
    PROJECT, '..', 'Finance Templates', '2026 TBs', 'Consolidated Accounts'))

# Months in load order. Filenames are the client's own.
MONTH_FILES = [
    ('2026-01', '2026-01-31', 'Jan 2026 Consolidated Accounts.xlsx'),
    ('2026-02', '2026-02-28', 'Feb 2026 Consolidated Accounts.xlsx'),
    ('2026-03', '2026-03-31', 'March 2026 Consolidated Accounts.xlsx'),
    ('2026-04', '2026-04-30', 'April 2026 Consolidated Accounts.xlsx'),
    ('2026-05', '2026-05-31', 'May 2026 Consolidated Accounts.xlsx'),
    ('2026-06', '2026-06-30', 'June 2026 Consolidated Accounts.xlsx'),
]

# tab -> (seed suffix, company_name used in the models, currency, synthetic-code prefix)
ENTITIES = {
    'ZAAC TB':   ('zaac',    'ZAAC',    'KES', 'ZAAC'),
    'ZARIB TB':  ('zarib',   'ZARIB',   'KES', 'ZARIB'),
    'Zamre TB':  ('zamre',   'ZAMRE',   'KES', 'ZAMRE'),
    'ZHL':       ('zhl',     'ZHL',     'KES', 'ZHL'),
    'C & P':     ('c_p',     'C&P',     'KES', 'CP'),
    'Rwanda':    ('rwanda',  'RWANDA',  'RWF', 'RW'),
    'Nigeria':   ('nigeria', 'NIGERIA', 'NGN', 'NG'),
    'Malawi TB': ('malawi',  'MALAWI',  'MWK', 'MW'),
    'MENA TB':   ('mena',    'MENA',    'AED', 'MENA'),
    'DRC TB':    ('drc',     'DRC',     'USD', 'DRC'),
    'ZATL':      ('zatl',    'ZATL',    'TZS', 'ZATL'),
}

STD_HEADER = ['Entry_No', 'Posting_Date', 'Document_Type', 'Document_No', 'External_Document_No',
    'Transaction_No', 'G_L_Account_No', 'Description', 'Amount', 'Debit_Amount', 'Credit_Amount',
    'Source_Code', 'Source_Type', 'Source_No', 'Dimension_Set_ID', 'Global_Dimension_1_Code',
    'Global_Dimension_2_Code', 'Business_Unit_Code', 'IC_Partner_Code', 'User_ID', 'Reversed',
    'Reversed_by_Entry_No', 'Reversed_Entry_No']
MENA_HEADER = ['Posting_Date', 'Description', 'Net_Debit_Credit', 'Amount_KES']
COA_HEADER = ['G_L_Account_No', 'Name', 'Account_Type', 'Account_Category', 'Income_Balance',
    'Debit_Credit', 'Direct_Posting', 'Blocked', 'Indentation', 'Totaling',
    'Consol_Translation_Method']

# Header-word detection. Mirrors load_tb.py; keep the two in step.
CODE_HDRS = {'a/c no', 'bc code', 'bc codes', 'account no', 'code'}
DESC_HDRS = {'description', 'account name', 'account'}
AMT_POST  = {'amount after accruals', 'amount after accrual', 'amount after accural', 'net amount'}
AMT_PLAIN = {'amount', 'dr', 'tzs', 'net debit /(credit)', 'net debit/(credit)', 'local', 'debit'}
CR_HDRS   = {'cr', 'credit'}
KES_HDRS  = {'kes'}
CAT_HDRS  = {'category', 'categories'}
# An 'Accruals' column alongside 'Amount'. Some tabs then carry a netted column
# with a proper header ('Net Amount', 'Amount after Accrual') and some leave that
# column with NO header at all — C&P is netted in June and unlabelled in the
# other months. Where the netted column is unlabelled we have to net it
# ourselves, because accrual-only rows have an EMPTY Amount and would otherwise
# be dropped: C&P's blank-coded 'Accrued Income' row is 11,077,486 of pure
# accrual, and losing it left both SFP Accrued Income and SCI Pension Admin Fee
# short by exactly that figure.
ACCRUAL_HDRS = {'accrual', 'accruals'}

SKIP_DESC = {'assets', 'non-current assets', 'current assets', 'fixed assets', 'liabilities',
    'equity', 'equity and liabilities', 'owners equity', 'current liabilities',
    'non-current liabilities', 'income', 'expenses', 'total', 'total assets', 'total equity',
    'total equity and liabilities', 'total income', 'total expenses', ''}

RATE_CURR = {'rwanda franc': 'RWF', 'malawi kwacha': 'MWK', 'ugx shs': 'UGX',
             'nigeria naira': 'NGN', 'tanzania shs': 'TZS', 'usd': 'USD', 'aed': 'AED'}

SCI_CATS = {'income', 'expenses', 'expense'}


def norm(s):
    return re.sub(r'\s+', ' ', str(s).replace('\xa0', ' ').strip().lower()) if s is not None else ''


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return None
    try:
        return float(str(v).replace(',', '').strip())
    except Exception:
        return None


def slug(desc, prefix, taken):
    """Readable, stable synthetic code: PREFIX-SLUG, numeric suffix only on collision."""
    body = re.sub(r'[^A-Z0-9]', '', str(desc).upper())[:18] or \
        hashlib.md5(str(desc).encode('utf-8')).hexdigest()[:8].upper()
    code = f'{prefix}-{body}'
    if code not in taken:
        return code
    for i in range(2, 99):
        alt = f'{code}{i}'
        if alt not in taken:
            return alt
    return f'{prefix}-{hashlib.md5(str(desc).encode("utf-8")).hexdigest()[:10].upper()}'


# ---------------------------------------------------------------- reading packs

def sheet_rows(path, wanted=None):
    """Materialise sheets once. ws.cell() random access on a read_only sheet is
    O(n^2); iter_rows is not. This is the difference between 35s and 0.2s a pack."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for name in wb.sheetnames:
        if wanted is not None and name not in wanted:
            continue
        out[name] = [r for r in wb[name].iter_rows(values_only=True)]
    wb.close()
    return out


def find_header(rows, maxscan=12):
    for i, row in enumerate(rows[:maxscan]):
        vals = {c: norm(v) for c, v in enumerate(row[:14])}
        if not any(v in DESC_HDRS for v in vals.values()):
            continue
        roles = {}
        for c, v in sorted(vals.items()):
            if v in CODE_HDRS and 'code' not in roles: roles['code'] = c
            elif v in DESC_HDRS and 'desc' not in roles: roles['desc'] = c
            elif v in KES_HDRS and 'kes' not in roles: roles['kes'] = c
            elif v in AMT_POST and 'amt_post' not in roles: roles['amt_post'] = c
            elif v in CR_HDRS and 'cr' not in roles: roles['cr'] = c
            elif v in ACCRUAL_HDRS and 'accrual' not in roles: roles['accrual'] = c
            elif v in AMT_PLAIN and 'amt_plain' not in roles: roles['amt_plain'] = c
            elif v in CAT_HDRS and 'cat' not in roles: roles['cat'] = c
        if 'desc' in roles and ('amt_post' in roles or 'amt_plain' in roles):
            return i, roles
    return None, None


def read_entity(rows):
    """-> (roles, [(source_code|None, description, local_amt, kes_amt|None, category|None)])

    Amounts are summed per (code, description) so a tab that lists the same
    account twice is added, not overwritten. 'Amount after Accrual' wins over
    'Amount' where the client supplies both, because the accrual column is what
    the client's own SCI/SFP is built from.
    """
    hi, roles = find_header(rows)
    if hi is None:
        return None, None
    ccode = roles.get('code'); cdesc = roles['desc']
    camt = roles['amt_post'] if 'amt_post' in roles else roles.get('amt_plain')
    ccr = roles.get('cr'); ckes = roles.get('kes'); ccat = roles.get('cat')
    # Net the accrual in ourselves only when the client has not already given us a
    # netted column; otherwise we would double-count it.
    cacc = roles.get('accrual') if 'amt_post' not in roles else None
    agg = collections.OrderedDict()
    for row in rows[hi + 1:]:
        def cell(c):
            return row[c] if c is not None and c < len(row) else None
        desc = cell(cdesc)
        nd = norm(desc)
        if nd in SKIP_DESC or nd.startswith('total') or nd.startswith('period'):
            continue
        loc = num(cell(camt))
        acc = num(cell(cacc)) if cacc is not None else None
        if acc is not None:
            loc = (loc or 0.0) + acc
        if ccr is not None:                      # Dr / Cr split layout: net debit-positive
            loc = (loc or 0.0) - (num(cell(ccr)) or 0.0)
        kes = num(cell(ckes)) if ckes is not None else None
        if loc is None and kes is None:
            continue
        code = None
        if ccode is not None and cell(ccode) not in (None, ''):
            code = str(cell(ccode)).strip()
        cat = str(cell(ccat)).strip() if ccat is not None and cell(ccat) else None
        key = (code or '', nd)
        if key not in agg:
            agg[key] = [code, str(desc).strip() if desc else '', 0.0,
                        0.0 if ckes is not None else None, cat]
        agg[key][2] += (loc or 0.0)
        if ckes is not None:
            agg[key][3] += (kes or 0.0)
    return roles, [tuple(v) for v in agg.values()]


def read_rates(rows, period):
    """-> {'CLOSING': {ccy: rate}, 'AVERAGE': {ccy: rate}} from the pack's Rates tab.

    Both are needed and they are different numbers. Surveying every per-entity
    formula on `KES consolidated TB` shows the client translates the balance sheet
    at the month's rate and the P&L at the period average: of the references whose
    statement is known, SFP uses the month row 242 times against 28 for the
    average, and SCI uses the average 226 times against 6. `int_fx_translation`
    already applies CLOSING to SFP and AVERAGE to SCI, so writing the month's rate
    into both — as the first cut of this script did — silently translated every
    P&L line at the wrong rate.

    The 'Average' row is a period-to-date average that grows each month, which is
    the right basis for marts holding cumulative year-to-date figures.
    """
    if rows is None:
        return None
    hdr = None; cols = {}
    for i, row in enumerate(rows[:12]):
        vals = {c: norm(v) for c, v in enumerate(row)}
        hit = {c: RATE_CURR[v] for c, v in vals.items() if v in RATE_CURR}
        if len(hit) >= 3:
            hdr = i; cols = hit; break
    if hdr is None:
        return None
    closing = average = None
    for row in rows[hdr + 1:]:
        a = row[0] if row else None
        if isinstance(a, datetime.datetime) and f'{a.year:04d}-{a.month:02d}' == period:
            closing = row
        elif isinstance(a, str) and 'average' in a.lower():
            average = row
    if closing is None:
        return None

    def pick(row):
        out = {'KES': 1.0}
        if row is None:
            return out
        for c, cur in cols.items():
            v = num(row[c]) if c < len(row) else None
            if v:
                out[cur] = v
        return out

    cl = pick(closing)
    av = pick(average)
    # Some packs (January) carry no usable Average row; fall back to closing and
    # let the report say so rather than silently inventing a number.
    for cur, v in cl.items():
        av.setdefault(cur, v)
    return {'CLOSING': cl, 'AVERAGE': av}


# ------------------------------------------------------------- codebook & rules

def snapshot_codebook(src_dir):
    """entity -> {'d2c': norm(description) -> [codes], 'c2d': code -> [norm descriptions]}

    Snapshotted from the seeds BEFORE they are rewritten (or from a backup via
    --codebook-dir). Two jobs:
      * resolve blank-code and description-only rows to the code the prior seed
        used, so an emptied seed does not lose them;
      * re-alias legacy local codes back onto BC codes. The final Consolidated
        Accounts packs code ZARIB / ZAMRE / ZHL / Malawi / Rwanda with legacy
        local numbers (7380/000, 1020/000, 6100) while account_map is keyed on
        BC codes, so taking the source code at face value would orphan most of
        those entities' mappings.
    gl_entry is read first because it reflects what actually loaded; gl_account
    fills in anything only present in the chart.
    """
    book = collections.defaultdict(lambda: {'d2c': collections.defaultdict(list),
                                            'c2d': collections.defaultdict(list)})
    for tab, (seed, company, cur, prefix) in ENTITIES.items():
        if seed == 'mena':
            continue
        for fn, dfield in ((f'gl_entry_{seed}.csv', 'Description'),
                           (f'gl_account_{seed}.csv', 'Name')):
            p = os.path.join(src_dir, fn)
            if not os.path.exists(p):
                continue
            for r in csv.DictReader(open(p, newline='')):
                d = norm(r.get(dfield)); c = (r.get('G_L_Account_No') or '').strip()
                if not (d and c):
                    continue
                if c not in book[seed]['d2c'][d]:
                    book[seed]['d2c'][d].append(c)
                if d not in book[seed]['c2d'][c]:
                    book[seed]['c2d'][c].append(d)
    return book


def load_rules():
    overrides = {}   # (seed, source_code, norm_desc) -> (code, basis, flag, note)
    p = os.path.join(HERE, 'code_overrides.csv')
    if os.path.exists(p):
        for r in csv.DictReader(open(p, newline='')):
            key = (r['entity'].strip(), (r['source_code'] or '').strip(),
                   norm(r['norm_description']))
            overrides[key] = (r['assigned_code'].strip(), r.get('basis', ''),
                              r.get('review_flag', ''), r.get('note', ''))
    bridge = {}      # norm_desc -> (code, basis, flag, note)   Nigeria only
    p = os.path.join(HERE, 'code_bridge_nigeria.csv')
    if os.path.exists(p):
        for r in csv.DictReader(open(p, newline='')):
            bridge[norm(r['norm_description'])] = (
                r['assigned_code'].strip(), r.get('basis', 'RAW_TB_BRIDGE'),
                r.get('review_flag', ''), r.get('note', ''))
    return overrides, bridge


def resolve_codes(observed, codebook, overrides, bridge):
    """observed: seed -> {(source_code, norm_desc): description}
    -> assign: seed -> {(source_code, norm_desc): (code, basis, flag, note)}"""
    assign = {}
    for tab, (seed, company, cur, prefix) in ENTITIES.items():
        if seed == 'mena':
            continue
        obs = observed.get(seed, {})
        # which descriptions share a source code
        by_code = collections.defaultdict(set)
        for (sc, nd) in obs:
            if sc:
                by_code[sc].add(nd)
        out = {}
        taken = set()
        # deterministic order so codes never depend on dict/row ordering
        for (sc, nd) in sorted(obs):
            key = (seed, sc, nd)
            if key in overrides:
                out[(sc, nd)] = overrides[key]
                taken.add(overrides[key][0])
                continue
            if seed == 'nigeria' and not sc and nd in bridge:
                out[(sc, nd)] = bridge[nd]
                taken.add(bridge[nd][0])
                continue
            cb = codebook.get(seed) or {'d2c': {}, 'c2d': {}}
            prior_for_desc = (cb['d2c'].get(nd) or [None])[0]
            prior_descs_for_code = cb['c2d'].get(sc) or []

            if sc and len(by_code[sc]) == 1:
                if sc in cb['c2d'] and nd in prior_descs_for_code:
                    out[(sc, nd)] = (sc, 'SOURCE_CODE_CONFIRMED', '',
                                     'Source code and description both match the prior seed')
                elif prior_for_desc and prior_for_desc != sc:
                    # legacy local code in the pack; prior seed keyed this same
                    # account on a BC code, which is what account_map expects
                    out[(sc, nd)] = (prior_for_desc, 'ALIAS_BY_DESCRIPTION', '',
                                     f'Pack codes this account {sc}; prior seed and account_map key it {prior_for_desc}')
                    taken.add(prior_for_desc)
                    continue
                elif prior_descs_for_code:
                    out[(sc, nd)] = (sc, 'CODE_DESC_MISMATCH', 'REVIEW',
                                     f'{sc} exists in the prior seed as {prior_descs_for_code[0]!r}, not {nd!r} — '
                                     'it may be inheriting the wrong account_map row')
                else:
                    out[(sc, nd)] = (sc, 'SOURCE_CODE_NEW', 'REVIEW',
                                     'Code not present in the prior seed; needs an account_map row')
                taken.add(sc)
                continue
            if sc:                                        # duplicated code
                existing = prior_for_desc
                if existing:
                    out[(sc, nd)] = (existing, 'CODEBOOK_DUP', 'REVIEW',
                                     f'{sc} reused by the client; prior seed keyed this account as {existing}')
                    taken.add(existing)
                else:
                    n = 2
                    while f'{sc}.{n}' in taken:
                        n += 1
                    code = f'{sc}.{n}'
                    out[(sc, nd)] = (code, 'DUP_SUFFIX', 'REVIEW',
                                     f'{sc} carries more than one description in the packs and no rule exists')
                    taken.add(code)
                continue
            existing = prior_for_desc                     # blank source code
            if existing:
                out[(sc, nd)] = (existing, 'CODEBOOK_BLANK', '',
                                 'Blank A/C No in source; matched to the prior seed by description')
                taken.add(existing)
            else:
                code = slug(obs[(sc, nd)], prefix, taken)
                out[(sc, nd)] = (code, 'SYNTHETIC', 'REVIEW',
                                 'Blank A/C No in source and no prior seed match; needs an account_map row')
                taken.add(code)

        # ---- collision guard -------------------------------------------------
        # Nothing above may hand the same assigned code to two different accounts:
        # that would silently merge balances the client keeps apart (e.g. ZHL
        # 'Bank Charges' and 'General expense', DRC's two petty-cash accounts).
        # Keep the best-evidenced description on the bare code, suffix the rest.
        rank = {'EXISTING_SEED': 0, 'RAW_TB_BRIDGE': 0, 'AMOUNT_MATCH+NAME': 0,
                'CLIENT_RECODE': 0, 'AMOUNT_MATCH': 1, 'BLANK_CODE_PINNED': 0,
                'DUP_SPLIT': 0, 'ALPHA_FIRST': 1, 'SOURCE_CODE_CONFIRMED': 1,
                'ALIAS_BY_DESCRIPTION': 2, 'SOURCE_CODE_NEW': 2,
                'CODE_DESC_MISMATCH': 3, 'CODEBOOK_DUP': 3, 'CODEBOOK_BLANK': 3,
                'DUP_SUFFIX': 4, 'SYNTHETIC': 5}
        by_assigned = collections.defaultdict(list)
        for k, v in out.items():
            by_assigned[v[0]].append(k)
        for code, keys in sorted(by_assigned.items()):
            # Same description arriving under two source codes is ONE account the
            # client re-coded mid-year, not two — those must stay merged.
            groups = collections.OrderedDict()
            for k in sorted(keys, key=lambda k: (rank.get(out[k][1], 9), k[1])):
                groups.setdefault(k[1], []).append(k)
            if len(groups) < 2:
                continue
            keys = [g[0] for g in groups.values()]
            for n, k in enumerate(keys[1:], start=2):
                while f'{code}.{n}' in taken:
                    n += 1
                newcode = f'{code}.{n}'
                _, basis, _, note = out[k]
                out[k] = (newcode, basis + '+COLLISION', 'REVIEW',
                          (note + ' | ' if note else '') +
                          f'Would have collided with {keys[0][1]!r} on {code}; separated to keep the accounts distinct')
                taken.add(newcode)
        assign[seed] = out
    return assign


# --------------------------------------------------------------------- writing

def write_gl_entry(path, header, rows_by_period, is_mena):
    """rows_by_period: ordered list of (pend, period, [(code, desc, mv, mv_kes)])"""
    with open(path, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=header)
        w.writeheader()
        entry = 0
        for pend, period, movements in rows_by_period:
            for code, desc, mv, mvk in sorted(movements, key=lambda x: (str(x[0]), str(x[1]))):
                entry += 1
                if is_mena:
                    w.writerow({'Posting_Date': pend, 'Description': desc,
                                'Net_Debit_Credit': round(mv, 2),
                                'Amount_KES': round(mvk or 0.0, 2)})
                else:
                    rec = {k: '' for k in header}
                    rec.update({'Entry_No': entry, 'Posting_Date': pend,
                                'Document_No': f'WB-{period}', 'Transaction_No': entry,
                                'G_L_Account_No': code, 'Description': desc,
                                'Amount': round(mv, 2),
                                'Debit_Amount': round(mv, 2) if mv > 0 else 0,
                                'Credit_Amount': round(-mv, 2) if mv < 0 else 0,
                                'Source_Code': 'WORKBOOK', 'Dimension_Set_ID': 0,
                                'User_ID': 'RESEED_CONSOLIDATED_ACCOUNTS', 'Reversed': 'false'})
                    w.writerow(rec)


def merge_gl_account(path, chart):
    """chart: code -> (name, category). Keeps existing rows, appends new codes."""
    rows = []
    seen = set()
    if os.path.exists(path):
        for r in csv.DictReader(open(path, newline='')):
            rows.append(r)
            seen.add((r.get('G_L_Account_No') or '').strip())
    added = []
    for code, (name, cat) in sorted(chart.items()):
        if code in seen:
            continue
        inc_bal = ''
        if cat:
            inc_bal = 'Income Statement' if norm(cat) in SCI_CATS else 'Balance Sheet'
        rec = {k: '' for k in COA_HEADER}
        rec.update({'G_L_Account_No': code, 'Name': name, 'Account_Type': 'Posting',
                    'Account_Category': cat or '', 'Income_Balance': inc_bal,
                    'Direct_Posting': 'true', 'Blocked': 'false', 'Indentation': 0})
        rows.append(rec); added.append(code)
    with open(path, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=COA_HEADER, extrasaction='ignore')
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, '') for k in COA_HEADER})
    return added


def write_fx(rates_by_period, implied):
    """Keep any period we did not reload; rewrite the ones we did.

    CLOSING comes from the period's own row on the pack's Rates tab, AVERAGE from
    that pack's 'Average' row (a period-to-date average). rate_source records
    which, so the choice stays auditable. The client's Rates tab cites CBK for
    most currencies and Oanda for Kwacha and Naira.
    """
    path = os.path.join(REF, 'fx_rate.csv')
    fld = ['currency', 'period', 'rate_type', 'rate_to_kes', 'rate_source']
    keep = []
    if os.path.exists(path):
        for r in csv.DictReader(open(path, newline='')):
            if r['period'] not in rates_by_period:
                keep.append(r)
            fld = list(r.keys())
    with open(path, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fld)
        w.writeheader()
        for r in keep:
            w.writerow(r)
        for period in sorted(rates_by_period):
            byte = rates_by_period[period]
            for rt in ('CLOSING', 'AVERAGE'):
                for cur, val in sorted(byte[rt].items()):
                    src = 'CBK' if cur == 'KES' else f'RATES_TAB_{rt}'
                    w.writerow({'currency': cur, 'period': period, 'rate_type': rt,
                                'rate_to_kes': f'{val:.10f}', 'rate_source': src})


# ------------------------------------------------------------------------ main

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--write', action='store_true', help='write the seeds (default is a dry run)')
    ap.add_argument('--packs', default=PACKS)
    ap.add_argument('--report-dir', default=HERE)
    ap.add_argument('--codebook-dir', default=BRONZE,
                    help='where to snapshot the prior seeds from (point at a backup if the '
                         'seeds have already been rewritten)')
    args = ap.parse_args()

    missing = [fn for _, _, fn in MONTH_FILES if not os.path.exists(os.path.join(args.packs, fn))]
    if missing:
        sys.exit('Missing pack(s) in %s:\n  %s' % (args.packs, '\n  '.join(missing)))

    codebook = snapshot_codebook(args.codebook_dir)   # BEFORE anything is rewritten
    overrides, bridge = load_rules()

    # ---- pass 1: read every pack -------------------------------------------
    packs = {}          # period -> seed -> [(code, desc, loc, kes, cat)]
    rates_by_period = {}
    for period, pend, fname in MONTH_FILES:
        rows = sheet_rows(os.path.join(args.packs, fname))
        packs[period] = {}
        for tab, (seed, company, cur, prefix) in ENTITIES.items():
            if tab not in rows:
                continue
            roles, rr = read_entity(rows[tab])
            if roles is None:
                print(f'  !! {period} {tab}: no header found — skipped')
                continue
            packs[period][seed] = rr
        r = read_rates(rows.get('Rates'), period)
        if r:
            rates_by_period[period] = r

    observed = collections.defaultdict(dict)
    for period in packs:
        for seed, rr in packs[period].items():
            if seed == 'mena':
                continue
            for code, desc, loc, kes, cat in rr:
                observed[seed][((code or ''), norm(desc))] = desc

    assign = resolve_codes(observed, codebook, overrides, bridge)

    # implied FX from the packs' own KES columns (median of KES/local on big rows)
    implied = {}
    for tab, (seed, company, cur, prefix) in ENTITIES.items():
        if cur == 'KES':
            continue
        for period in packs:
            rr = packs[period].get(seed) or []
            ratios = [k / loc for code, desc, loc, k, cat in rr
                      if k is not None and loc and abs(loc) > 1000]
            if len(ratios) >= 5:
                ratios.sort()
                implied[(cur, period)] = ratios[len(ratios) // 2]

    # ---- pass 2: movements, in month order ---------------------------------
    audit = []
    per_seed = {}
    recon = []
    for tab, (seed, company, cur, prefix) in ENTITIES.items():
        is_mena = (seed == 'mena')
        cum = collections.defaultdict(float)
        cum_kes = collections.defaultdict(float)
        desc_of = {}
        by_period = []
        chart = {}
        for period, pend, fname in MONTH_FILES:
            rr = packs[period].get(seed)
            if rr is None:
                by_period.append((pend, period, []))
                continue
            ytd = collections.defaultdict(float)
            ytd_kes = collections.defaultdict(float)
            src_total = 0.0
            for code, desc, loc, kes, cat in rr:
                if is_mena:
                    key = desc                                  # raw string is the key
                else:
                    code_out, basis, flag, note = assign[seed][((code or ''), norm(desc))]
                    key = code_out
                    chart[code_out] = (desc, cat)
                    audit.append({'entity': company, 'period': period,
                                  'source_code': code or '', 'description': desc,
                                  'assigned_code': code_out, 'basis': basis,
                                  'review_flag': flag, 'note': note,
                                  'ytd_local': round(loc, 2)})
                desc_of[key] = desc
                ytd[key] += loc
                src_total += loc
                if kes is not None:
                    ytd_kes[key] += kes
            # MENA needs Amount_KES; Jan/Feb packs have no KES column, so translate
            # at the month's closing rate. NB rates_by_period is keyed by rate
            # type ('CLOSING' / 'AVERAGE') — reading it as a flat currency map
            # silently yielded no rate and wrote Amount_KES = 0 for those months.
            if is_mena and not ytd_kes:
                rate = ((rates_by_period.get(period) or {}).get('CLOSING') or {}).get(cur)
                if rate:
                    for k, v in ytd.items():
                        ytd_kes[k] = v * rate
            movements = []
            for key in set(ytd) | set(cum):
                mv = round(ytd.get(key, 0.0) - cum[key], 2)
                mvk = round(ytd_kes.get(key, 0.0) - cum_kes[key], 2) if is_mena else None
                if abs(mv) < 0.005 and (not is_mena or abs(mvk) < 0.005):
                    continue
                movements.append((key, desc_of.get(key, ''), mv, mvk))
            by_period.append((pend, period, movements))
            for key in set(ytd) | set(cum):
                cum[key] = ytd.get(key, 0.0)
                if is_mena:
                    cum_kes[key] = ytd_kes.get(key, 0.0)
            recon.append({'entity': company, 'period': period,
                          'source_accounts': len(rr),
                          'seed_accounts': len([k for k in cum if abs(cum[k]) > 0.005]),
                          'source_ytd_total': round(src_total, 2),
                          'seed_cumulative_total': round(sum(cum.values()), 2),
                          'difference': round(sum(cum.values()) - src_total, 2),
                          'movement_rows': len(movements)})
        per_seed[seed] = (by_period, chart, is_mena)

    # ---- report -------------------------------------------------------------
    os.makedirs(args.report_dir, exist_ok=True)
    with open(os.path.join(args.report_dir, 'reseed_audit.csv'), 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['entity', 'period', 'source_code', 'description',
            'assigned_code', 'basis', 'review_flag', 'note', 'ytd_local'])
        w.writeheader(); w.writerows(audit)
    with open(os.path.join(args.report_dir, 'reseed_recon.csv'), 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['entity', 'period', 'source_accounts', 'seed_accounts',
            'source_ytd_total', 'seed_cumulative_total', 'difference', 'movement_rows'])
        w.writeheader(); w.writerows(recon)
    json.dump({'implied_fx': {f'{c}|{p}': v for (c, p), v in implied.items()},
               'rates_tab': rates_by_period},
              open(os.path.join(args.report_dir, 'reseed_fx.json'), 'w'), indent=1)

    basis_count = collections.Counter((a['entity'], a['basis']) for a in audit)
    print('\nCode resolution by entity and basis (account-months):')
    for (ent, basis), n in sorted(basis_count.items()):
        print(f'  {ent:8} {basis:16} {n:6}')
    flagged = {(a['entity'], a['assigned_code'], a['description']) for a in audit if a['review_flag']}
    print(f'\nAccounts flagged for review: {len(flagged)}')
    worst = sorted(recon, key=lambda r: -abs(r['difference']))[:6]
    print('\nLargest source-vs-seed differences (should all be ~0):')
    for r in worst:
        print(f"  {r['entity']:8} {r['period']}  diff {r['difference']:>16,.2f}")

    if not args.write:
        print('\nDRY RUN — nothing written. Reports in %s. Re-run with --write.' % args.report_dir)
        return

    for tab, (seed, company, cur, prefix) in ENTITIES.items():
        by_period, chart, is_mena = per_seed[seed]
        path = os.path.join(BRONZE, f'gl_entry_{seed}.csv')
        write_gl_entry(path, MENA_HEADER if is_mena else STD_HEADER, by_period, is_mena)
        if not is_mena:
            added = merge_gl_account(os.path.join(BRONZE, f'gl_account_{seed}.csv'), chart)
            if added:
                print(f'  {company}: {len(added)} new gl_account rows')
    write_fx(rates_by_period, implied)
    print('\nWROTE seeds. Next: dbt seed --full-refresh && dbt build, then reconcile.')


if __name__ == '__main__':
    main()
