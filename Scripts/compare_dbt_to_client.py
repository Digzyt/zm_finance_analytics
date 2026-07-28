#!/usr/bin/env python3
"""Compare the dbt subsidiary SCI/SFP marts to the client's own statements,
line by line, entity by entity, month by month — and write the recon workbook.

    python Scripts/extract_client_statements.py            # 1. the client side
    dbt seed --full-refresh && dbt build                   # 2. our side
    python Scripts/compare_dbt_to_client.py                # 3. the recon

Where each side comes from
--------------------------
**Client**  `Scripts/out/client_by_line.csv`, written by
`extract_client_statements.py` straight out of the `SCI Detailed` / `SFP Detailed`
tabs. Pass `--refresh` to re-run the extract first.

**Ours**  `subsidiary.rpt_subsidiary_sci` and `subsidiary.rpt_subsidiary_sfp`,
read either from Postgres (default) or from CSV exports:

    --marts postgres                     # reads dbt profiles.yml + $PG_PASSWORD
    --marts csv --marts-dir ../Outputs   # reads rpt_subsidiary_*.csv exports

The CSV route exists because the mart exports are the only thing that travels: it
takes `rpt_subsidiary_sci*.csv` / `rpt_subsidiary_sfp*.csv` from a folder, in
UTF-8, UTF-16 or with the NUL bytes Power BI's export leaves behind.

Sign basis
----------
Both sides are compared on the **statement-presentation basis** — income and
equity positive — because that is what the marts hold. The client's own
debit-positive figure is carried alongside in the workbook so any row can be
tied back to their pack by eye.

Reading a difference
--------------------
Every differing line is bucketed, and buckets are the point of the exercise:

    rounding / FX          under 0.5% and under KES 2m — translation noise
    we show nothing        the client reports a figure on this line, we report none
    client shows nothing   we report a figure they do not
    classification         both report, on a different line — a mapping question
    ... offset by X        another line in the same entity-month is out by the
                           mirror amount, so it is a reclassification between the
                           two, not a missing balance

The offset detection is what makes the difference list actionable: a pair that
offsets needs one mapping decision, whereas a difference with no partner means
something is genuinely absent from the seeds.
"""
import argparse
import collections
import csv
import datetime
import glob
import os
import re
import subprocess
import sys

SCRIPTS = os.path.dirname(os.path.abspath(__file__))
DATAMODEL = os.path.dirname(SCRIPTS)
ROOT = os.path.dirname(DATAMODEL)
SEEDS = os.path.join(DATAMODEL, 'seeds')
OUTDIR = os.path.join(SCRIPTS, 'out')
INTERNAL = os.path.join(ROOT, 'Internal')

MARTS = {'SCI': 'subsidiary.rpt_subsidiary_sci', 'SFP': 'subsidiary.rpt_subsidiary_sfp'}

TOL = 1.0                 # KES — inside this a line is a tie
FX_PCT, FX_ABS = 0.005, 2_000_000.0   # what counts as translation noise

ENTITY_ORDER = ['ZAAC', 'ZAMRE', 'ZARIB', 'MALAWI', 'MENA', 'NIGERIA', 'RWANDA',
                'ZHL', 'ZATL', 'DRC', 'C&P']


# --------------------------------------------------------------------- reference
def load_statement_lines():
    out = {}
    with open(os.path.join(SEEDS, 'reference', 'statement_line.csv'),
              newline='', encoding='utf-8-sig') as f:
        for r in csv.DictReader(f):
            out[r['statement_line_code']] = r
    return out


# ------------------------------------------------------------------ client side
def load_client(path):
    """(entity, period, statement, code) -> presentation-basis KES."""
    if not os.path.exists(path):
        sys.exit(f'client extract not found: {path}\n'
                 f'run: python Scripts/extract_client_statements.py')
    client, raw = {}, {}
    with open(path, newline='', encoding='utf-8-sig') as f:
        for r in csv.DictReader(f):
            k = (r['entity'], r['period'], r['statement'], r['statement_line_code'])
            client[k] = client.get(k, 0.0) + float(r['client_presentation_kes'])
            raw[k] = raw.get(k, 0.0) + float(r['client_basis_kes'])
    return client, raw


# --------------------------------------------------------------------- our side
def read_profiles(target=None):
    """Postgres connection settings out of dbt's profiles.yml.

    Looked for in ~/.dbt/profiles.yml first — that is the real one on Donald's
    machine; the copy in the repo is a template with placeholder credentials.
    `{{ env_var('X') }}` is resolved from the environment.
    """
    import yaml
    candidates = [os.path.expanduser('~/.dbt/profiles.yml'),
                  os.path.join(DATAMODEL, 'profiles.yml')]
    for path in candidates:
        if not os.path.exists(path):
            continue
        with open(path) as f:
            text = f.read()

        def env(m):
            name = m.group(1)
            default = m.group(2)
            return os.environ.get(name, default if default is not None else '')

        text = re.sub(r"\{\{\s*env_var\(\s*'([^']+)'\s*(?:,\s*'([^']*)')?\s*\)\s*\}\}",
                      env, text)
        cfg = yaml.safe_load(text)
        prof = cfg.get('zamara_finance') or next(iter(cfg.values()))
        tgt = target or prof.get('target')
        out = prof['outputs'].get(tgt)
        if out and out.get('type') == 'postgres':
            out['_profiles_path'] = path
            return out
    sys.exit('no postgres target found in profiles.yml — pass --marts csv instead')


def load_marts_postgres(target=None, dsn=None):
    import psycopg2
    if dsn:
        conn = psycopg2.connect(dsn)
        where = dsn
    else:
        p = read_profiles(target)
        if not p.get('password'):
            print('  note: no password resolved — set PG_PASSWORD', file=sys.stderr)
        conn = psycopg2.connect(host=p['host'], port=p.get('port', 5432),
                                user=p['user'], password=p.get('password'),
                                dbname=p['dbname'])
        where = f"{p['user']}@{p['host']}/{p['dbname']} (target from {p['_profiles_path']})"
    ours = collections.defaultdict(float)
    cur = conn.cursor()
    for which, table in MARTS.items():
        cur.execute(f'select company_name, period, statement_line_code, '
                    f'sum(amount_kes) from {table} group by 1,2,3')
        for co, period, code, amt in cur.fetchall():
            ours[(co, period, which, code)] += float(amt or 0)
    conn.close()
    return ours, where


def _read_csv_any(path):
    """Mart exports arrive UTF-8, UTF-16, or with NUL bytes. Read all three."""
    with open(path, 'rb') as f:
        data = f.read()
    for enc in ('utf-8-sig', 'utf-16', 'utf-16-le', 'latin-1'):
        try:
            text = data.decode(enc)
        except (UnicodeDecodeError, UnicodeError):
            continue
        text = text.replace('\x00', '')
        if text.count(',') > 2:
            return list(csv.DictReader(text.splitlines()))
    sys.exit(f'could not decode {path}')


def load_marts_csv(folder):
    ours = collections.defaultdict(float)
    found = []
    for which, table in MARTS.items():
        stem = table.split('.')[-1]
        files = sorted(glob.glob(os.path.join(folder, stem + '*.csv')))
        if not files:
            sys.exit(f'no {stem}*.csv in {folder}')
        for path in files:
            for r in _read_csv_any(path):
                r = {k.strip().lower().replace('﻿', ''): v for k, v in r.items()
                     if k}
                co = r.get('company_name')
                period = r.get('period')
                code = r.get('statement_line_code')
                amt = r.get('amount_kes') or '0'
                if not co or not code:
                    continue
                try:
                    val = float(str(amt).replace(',', '').strip() or 0)
                except ValueError:
                    continue
                ours[(co, period, which, code)] += val
            found.append(os.path.basename(path))
    return ours, ', '.join(found)


# ------------------------------------------------------------------- comparison
def classify(ours, client, offsets):
    d = ours - client
    if abs(d) <= TOL:
        return 'tie', ''
    if abs(client) <= TOL:
        return 'client shows nothing', ''
    if abs(ours) <= TOL:
        return 'we show nothing', ''
    if abs(d) < FX_ABS and abs(d) <= FX_PCT * abs(client):
        return 'rounding / FX', ''
    partner = offsets.get(round(-d, 0))
    if partner:
        return 'classification', partner
    return 'classification', ''


def compare(client, client_raw, ours, sl):
    rows, summary = [], []
    keys = set(client) | set(ours)
    groups = collections.defaultdict(list)
    for co, period, which, code in keys:
        groups[(co, period, which)].append(code)

    for (co, period, which), codes in sorted(groups.items()):
        # differences within this entity-month, for offset detection
        diffs = {}
        for code in codes:
            d = ours.get((co, period, which, code), 0.0) - client.get((co, period, which, code), 0.0)
            if abs(d) > TOL:
                diffs[code] = d
        offsets = {}
        for code, d in diffs.items():
            for other, od in diffs.items():
                if other == code:
                    continue
                if abs(od + d) <= max(TOL, 0.01 * abs(d)):
                    offsets.setdefault(round(od, 0),
                                       f'offset by {sl.get(other, {}).get("line_label", other)}')
        ties = offs = 0
        worst, absdiff = 0.0, 0.0
        for code in sorted(codes, key=lambda c: int(sl.get(c, {}).get('line_order', 9999))):
            o = ours.get((co, period, which, code), 0.0)
            cl = client.get((co, period, which, code), 0.0)
            craw = client_raw.get((co, period, which, code), 0.0)
            d = o - cl
            bucket, note = classify(o, cl, offsets)
            if bucket == 'tie':
                ties += 1
            else:
                offs += 1
                worst = max(worst, abs(d))
                absdiff += abs(d)
            meta = sl.get(code, {})
            rows.append({
                'entity': co, 'period': period, 'statement': which,
                'category_l1': meta.get('category_l1', ''),
                'category_l2': meta.get('category_l2', ''),
                'statement_line_code': code,
                'line_label': meta.get('line_label', code),
                'line_order': int(meta.get('line_order', 9999)),
                'dbt_kes': round(o, 2),
                'client_kes': round(cl, 2),
                'difference': round(d, 2),
                'difference_pct': (round(100 * d / cl, 2) if abs(cl) > TOL else ''),
                'status': 'tie' if bucket == 'tie' else 'DIFFERENCE',
                'reason': bucket if bucket != 'tie' else '',
                'note': note,
                'client_as_stated_kes': round(craw, 2)})

        def band(statement, cat, mult=None):
            t = 0.0
            for code, meta in sl.items():
                if meta['statement_type'] != statement or meta['category_l1'] != cat:
                    continue
                if mult is not None and float(meta['sign_multiplier']) != mult:
                    continue
                t += ours.get((co, period, which, code), 0.0)
            return t

        def bandc(statement, cat):
            return sum(client.get((co, period, which, code), 0.0)
                       for code, meta in sl.items()
                       if meta['statement_type'] == statement and meta['category_l1'] == cat)

        summary.append({
            'entity': co, 'period': period, 'statement': which,
            'lines_compared': len(codes), 'lines_tying': ties, 'lines_off': offs,
            'pct_tying': round(100 * ties / len(codes), 1) if codes else 0,
            'total_absolute_difference': round(absdiff, 2),
            'largest_difference': round(worst, 2),
            'dbt_income_or_assets': round(band(which, 'INCOME' if which == 'SCI' else 'ASSETS'), 2),
            'client_income_or_assets': round(bandc(which, 'INCOME' if which == 'SCI' else 'ASSETS'), 2),
            'dbt_expense_or_eandl': round(band(which, 'EXPENSES' if which == 'SCI' else 'EQUITY AND LIABILITIES'), 2),
            'client_expense_or_eandl': round(bandc(which, 'EXPENSES' if which == 'SCI' else 'EQUITY AND LIABILITIES'), 2),
        })
    for s in summary:
        s['income_or_assets_difference'] = round(
            s['dbt_income_or_assets'] - s['client_income_or_assets'], 2)
        s['expense_or_eandl_difference'] = round(
            s['dbt_expense_or_eandl'] - s['client_expense_or_eandl'], 2)
    rows.sort(key=lambda r: (r['entity'], r['period'], r['statement'], r['line_order']))
    return rows, summary


# ----------------------------------------------------------------------- output
def write_workbook(path, rows, summary, sl, source_note, periods):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HDR = PatternFill('solid', fgColor='1F3864')
    OKF = PatternFill('solid', fgColor='E2EFDA')
    BADF = PatternFill('solid', fgColor='FCE4D6')
    HOTF = PatternFill('solid', fgColor='F8CBAD')
    WHITE = Font(color='FFFFFF', bold=True, size=10)
    NUM = '#,##0;[Red](#,##0)'
    PCT = '0.0'

    wb = Workbook()

    def header(ws, cols):
        for i, c in enumerate(cols, 1):
            cell = ws.cell(1, i, c)
            cell.fill = HDR
            cell.font = WHITE
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
        ws.freeze_panes = ws.cell(2, 1)
        ws.row_dimensions[1].height = 30

    def table(ws, data, cols, widths=None, num_cols=(), highlight=None):
        header(ws, [c.replace('_', ' ').title() for c in cols])
        for i, r in enumerate(data, 2):
            for j, c in enumerate(cols, 1):
                cell = ws.cell(i, j, r.get(c))
                if c in num_cols:
                    cell.number_format = NUM
                if c.endswith('_pct') or c == 'pct_tying':
                    cell.number_format = PCT
            if highlight:
                highlight(ws, i, r, len(cols))
        for j, c in enumerate(cols, 1):
            w = (widths or {}).get(c)
            ws.column_dimensions[get_column_letter(j)].width = w or (
                26 if ('label' in c or 'code' in c or 'note' in c or 'reason' in c) else 15)
        ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}{len(data) + 1}'

    total_lines = len(rows)
    tying = sum(1 for r in rows if r['status'] == 'tie')
    absdiff = sum(abs(r['difference']) for r in rows)
    reasons = collections.Counter(r['reason'] for r in rows if r['reason'])

    # ----------------------------------------------------------------- Read Me
    ws = wb.active
    ws.title = 'Read Me'
    ws.column_dimensions['A'].width = 116
    body = [
        ('SCI and SFP: dbt vs the client, every entity, every month', 'title'),
        ('', ''),
        ('Headline', 'h'),
        (f'{tying:,} of {total_lines:,} line cells tie exactly '
         f'({100 * tying / total_lines:.1f}%).', 'big'),
        (f'Total absolute difference: KES {absdiff:,.0f} across '
         f'{total_lines - tying:,} differing cells.', ''),
        ('', ''),
        ('Where the two sides come from', 'h'),
        ('Client   the `SCI Detailed` / `SFP Detailed` tabs of the monthly '
         'Consolidated Accounts packs,', ''),
        ('         extracted by Scripts/extract_client_statements.py. Their '
         'statement as they state it.', ''),
        (f'Ours     subsidiary.rpt_subsidiary_sci / _sfp — {source_note}', ''),
        ('', ''),
        ('Sign basis', 'h'),
        ('Both sides are shown on the statement-presentation basis: income and '
         'equity positive.', ''),
        ('`Client As Stated Kes` carries the client\'s own debit-positive figure so '
         'a row can be tied', ''),
        ('back to their pack by eye.', ''),
        ('', ''),
        ('How to read a difference', 'h'),
    ]
    for reason, n in reasons.most_common():
        expl = {
            'rounding / FX': 'under 0.5% and under KES 2m — translation noise, not a mapping question',
            'we show nothing': 'the client reports a figure on this line and we report none — '
                               'usually an unmapped or missing account',
            'client shows nothing': 'we report a figure they do not — usually a line we have '
                                    'mapped somewhere they do not use',
            'classification': 'both report, on different lines — a mapping decision. '
                              'Where the Note says "offset by", the mirror amount sits on '
                              'that line, so it is a reclassification between the two',
        }.get(reason, '')
        body.append((f'    {reason:22s} {n:5,d} cells   {expl}', 'code'))
    body += [
        ('', ''),
        ('Tabs', 'h'),
        ('Summary            entity x month x statement: how many lines tie, and the '
         'income / expense /', ''),
        ('                   asset / equity totals side by side.', ''),
        ('Side by Side       every line cell, ours and theirs, with a reason on each '
         'difference.', ''),
        ('Differences Only   the same rows filtered to differences, largest first — '
         'the worklist.', ''),
        ('By Statement Line  which lines carry the difference, across all entities.', ''),
        ('Entity x Month     total absolute difference per entity per month.', ''),
        ('', ''),
        ('Scope', 'h'),
        (f'Periods: {", ".join(periods)}. January is absent by design — that pack has no '
         'per-entity', ''),
        ('Detailed tabs, only the consolidated group statement, so there is nothing to '
         'compare per entity.', ''),
        ('', ''),
        ('Re-running it', 'h'),
        ('    python Scripts/extract_client_statements.py', 'code'),
        ('    dbt seed --full-refresh && dbt build', 'code'),
        ('    python Scripts/compare_dbt_to_client.py', 'code'),
        ('', ''),
        (f'Generated {datetime.datetime.now():%Y-%m-%d %H:%M} by '
         f'Scripts/compare_dbt_to_client.py', 'small'),
    ]
    r = 1
    for text, kind in body:
        cell = ws.cell(r, 1, text)
        if kind == 'title':
            cell.font = Font(bold=True, size=14, color='1F3864')
        elif kind == 'h':
            cell.font = Font(bold=True, size=11, color='1F3864')
        elif kind == 'big':
            cell.font = Font(bold=True, size=12)
        elif kind == 'code':
            cell.font = Font(name='Consolas', size=9)
        elif kind == 'small':
            cell.font = Font(italic=True, size=9, color='808080')
        else:
            cell.font = Font(size=10)
        r += 1

    # ----------------------------------------------------------------- Summary
    ws = wb.create_sheet('Summary')
    scols = ['entity', 'period', 'statement', 'lines_compared', 'lines_tying',
             'lines_off', 'pct_tying', 'total_absolute_difference', 'largest_difference',
             'dbt_income_or_assets', 'client_income_or_assets', 'income_or_assets_difference',
             'dbt_expense_or_eandl', 'client_expense_or_eandl', 'expense_or_eandl_difference']
    numeric = set(scols[7:])

    def hl_summary(ws, i, r, ncols):
        if r['lines_off'] == 0:
            ws.cell(i, 6).fill = OKF
        elif r['total_absolute_difference'] > 10_000_000:
            ws.cell(i, 8).fill = HOTF
        else:
            ws.cell(i, 8).fill = BADF
    table(ws, summary, scols, num_cols=numeric, highlight=hl_summary,
          widths={'entity': 10, 'period': 9, 'statement': 10})

    # ------------------------------------------------------------ Side by Side
    dcols = ['entity', 'period', 'statement', 'category_l1', 'statement_line_code',
             'line_label', 'dbt_kes', 'client_kes', 'difference', 'difference_pct',
             'status', 'reason', 'note', 'client_as_stated_kes']
    dnum = {'dbt_kes', 'client_kes', 'difference', 'client_as_stated_kes'}

    def hl_diff(ws, i, r, ncols):
        if r['status'] != 'tie':
            ws.cell(i, 9).fill = HOTF if abs(r['difference']) > 10_000_000 else BADF
    ws = wb.create_sheet('Side by Side')
    table(ws, rows, dcols, num_cols=dnum, highlight=hl_diff,
          widths={'entity': 10, 'period': 9, 'statement': 10, 'category_l1': 20,
                  'line_label': 34, 'reason': 20, 'note': 34})

    ws = wb.create_sheet('Differences Only')
    diffs = sorted([r for r in rows if r['status'] != 'tie'],
                   key=lambda r: -abs(r['difference']))
    table(ws, diffs, dcols, num_cols=dnum, highlight=hl_diff,
          widths={'entity': 10, 'period': 9, 'statement': 10, 'category_l1': 20,
                  'line_label': 34, 'reason': 20, 'note': 34})

    # -------------------------------------------------------- By Statement Line
    per_line = collections.defaultdict(lambda: {'cells': 0, 'off': 0, 'abs': 0.0,
                                                'signed': 0.0, 'entities': set()})
    for r in rows:
        k = (r['statement'], r['line_order'], r['statement_line_code'], r['line_label'])
        s = per_line[k]
        s['cells'] += 1
        if r['status'] != 'tie':
            s['off'] += 1
            s['abs'] += abs(r['difference'])
            s['signed'] += r['difference']
            s['entities'].add(r['entity'])
    line_rows = [{'statement': k[0], 'statement_line_code': k[2], 'line_label': k[3],
                  'cells_compared': v['cells'], 'cells_off': v['off'],
                  'total_absolute_difference': round(v['abs'], 2),
                  'net_difference': round(v['signed'], 2),
                  'entities_affected': ', '.join(sorted(v['entities']))}
                 for k, v in sorted(per_line.items(), key=lambda kv: -kv[1]['abs'])]
    ws = wb.create_sheet('By Statement Line')
    table(ws, line_rows, list(line_rows[0].keys()),
          num_cols={'total_absolute_difference', 'net_difference'},
          widths={'line_label': 34, 'entities_affected': 46})

    # ---------------------------------------------------------- Entity x Month
    ws = wb.create_sheet('Entity x Month')
    ents = [e for e in ENTITY_ORDER if any(r['entity'] == e for r in rows)]
    ents += sorted({r['entity'] for r in rows} - set(ents))
    header(ws, ['Entity', 'Statement'] + list(periods) + ['Total', 'Lines off'])
    grid = collections.defaultdict(float)
    offgrid = collections.defaultdict(int)
    for r in rows:
        if r['status'] != 'tie':
            grid[(r['entity'], r['statement'], r['period'])] += abs(r['difference'])
            offgrid[(r['entity'], r['statement'])] += 1
    i = 2
    for e in ents:
        for which in ('SCI', 'SFP'):
            ws.cell(i, 1, e)
            ws.cell(i, 2, which)
            tot = 0.0
            for j, p in enumerate(periods):
                v = grid.get((e, which, p), 0.0)
                tot += v
                c = ws.cell(i, 3 + j, round(v, 0))
                c.number_format = NUM
                c.fill = OKF if v <= TOL else (HOTF if v > 10_000_000 else BADF)
            c = ws.cell(i, 3 + len(periods), round(tot, 0))
            c.number_format = NUM
            c.font = Font(bold=True)
            ws.cell(i, 4 + len(periods), offgrid.get((e, which), 0))
            i += 1
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 11
    for j in range(len(periods) + 2):
        ws.column_dimensions[get_column_letter(3 + j)].width = 15

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path


def write_csvs(outdir, rows, summary):
    os.makedirs(outdir, exist_ok=True)
    for name, data in (('recon_side_by_side', rows), ('recon_summary', summary)):
        p = os.path.join(outdir, name + '.csv')
        with open(p, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=list(data[0].keys()))
            w.writeheader()
            w.writerows(data)
        print(f'  wrote {p} ({len(data)} rows)')


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--marts', choices=['postgres', 'csv'], default='postgres',
                    help='where to read rpt_subsidiary_sci/sfp from')
    ap.add_argument('--marts-dir', default=os.path.join(ROOT, 'Outputs'),
                    help='folder of mart CSV exports, for --marts csv')
    ap.add_argument('--target', help='dbt profiles.yml target (default: its own default)')
    ap.add_argument('--dsn', help='postgres connection string, overrides profiles.yml')
    ap.add_argument('--label', help='how to describe the dbt source in the workbook '
                                    '(default: the connection itself)')
    ap.add_argument('--client', default=os.path.join(OUTDIR, 'client_by_line.csv'))
    ap.add_argument('--refresh', action='store_true',
                    help='re-run extract_client_statements.py first')
    ap.add_argument('--xlsx', default=os.path.join(
        INTERNAL, 'Phase1_SCI_SFP_Recon_dbt_vs_Client.xlsx'))
    ap.add_argument('--outdir', default=OUTDIR)
    a = ap.parse_args()

    if a.refresh:
        print('refreshing the client extract...')
        subprocess.check_call([sys.executable,
                               os.path.join(SCRIPTS, 'extract_client_statements.py'),
                               '--quiet'])

    sl = load_statement_lines()
    client, client_raw = load_client(a.client)
    if a.marts == 'csv':
        ours, source_note = load_marts_csv(a.marts_dir)
        source_note = f'CSV exports in {a.marts_dir} ({source_note})'
    else:
        ours, source_note = load_marts_postgres(a.target, a.dsn)
        source_note = f'Postgres — {source_note}'
    if a.label:
        source_note = a.label
    print(f'client cells: {len(client):,}   dbt cells: {len(ours):,}')

    # Only compare months the client actually publishes per entity.
    periods = sorted({k[1] for k in client})
    ours = {k: v for k, v in ours.items() if k[1] in periods}

    rows, summary = compare(client, client_raw, ours, sl)
    write_csvs(a.outdir, rows, summary)
    path = write_workbook(a.xlsx, rows, summary, sl, source_note, periods)
    print(f'  wrote {path}')

    tying = sum(1 for r in rows if r['status'] == 'tie')
    print()
    print(f'line cells compared : {len(rows):,}')
    print(f'tying               : {tying:,} ({100 * tying / len(rows):.1f}%)')
    print(f'total abs difference: KES {sum(abs(r["difference"]) for r in rows):,.0f}')
    print()
    print('by reason:')
    for reason, n in collections.Counter(r['reason'] for r in rows if r['reason']).most_common():
        amt = sum(abs(r['difference']) for r in rows if r['reason'] == reason)
        print(f'   {reason:22s} {n:5,d} cells   KES {amt:>18,.0f}')
    print()
    print('%-8s %-6s %6s %6s %6s %20s' % ('entity', 'stmt', 'lines', 'tie', 'off', 'abs difference'))
    agg = collections.defaultdict(lambda: [0, 0, 0.0])
    for r in rows:
        a2 = agg[(r['entity'], r['statement'])]
        a2[0] += 1
        a2[1] += 1 if r['status'] == 'tie' else 0
        a2[2] += abs(r['difference'])
    for (e, w), (n, t, d) in sorted(agg.items(), key=lambda kv: -kv[1][2]):
        print('%-8s %-6s %6d %6d %6d %20s' % (e, w, n, t, n - t, format(d, ',.0f')))


if __name__ == '__main__':
    main()
