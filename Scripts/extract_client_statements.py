#!/usr/bin/env python3
"""Extract the client's own SCI Detailed / SFP Detailed, per entity per month,
out of the `Consolidated Accounts` packs and into one spreadsheet.

Why this exists
---------------
The client's `SCI Detailed` and `SFP Detailed` tabs ARE their statements: one
column per entity, one row per line, and every figure a formula terminating in
an entity TB cell. They are the thing our marts have to agree with, so they need
to be readable on their own — as a flat table, for every month, without opening
six workbooks.

This script only reads. It does not touch seeds, mappings or models. Point it at
the pack folder and it rebuilds the extract from scratch, so next month is one
command: drop `July 2026 Consolidated Accounts.xlsx` into the folder and re-run.

    python Scripts/extract_client_statements.py                 # writes the xlsx + CSVs
    python Scripts/extract_client_statements.py --period 2026-07 # one month only
    python Scripts/extract_client_statements.py --quiet

Outputs
-------
    Internal/Phase1_Client_SCI_SFP_Extract.xlsx   the reviewable workbook
    Scripts/out/client_statements_long.csv        one row per period/entity/line
    Scripts/out/client_by_line.csv                aggregated to statement_line_code

`compare_dbt_to_client.py` reads the two CSVs — run this first.

Three things the reader has to get right
---------------------------------------
1. **Entity columns move between tabs and between packs.** `SCI Detailed` starts
   at column C, `SFP Detailed` at column D, and the entity order is not stable.
   Every tab's own header row is read: rows 1-8 are scanned and the row with the
   most recognised entity names wins. Taking first-seen-wins instead picks up the
   sheet title in A1 ("Zamara Holdings Limited") and reads it as ZHL's column.
2. **The SCI expense section is group headers over account-level detail rows.**
   Column A carries the client's marker (`P&L` on the SCI, `Balance Sheet` on the
   SFP) and is blank on group headers and subtotals. An expense detail row's own
   label is an account name ("Emol.Pack-Salaries"), not a statement line, so it
   inherits its group ("Personnel Costs"). Testing for a non-empty marker is the
   structural rule; matching the marker text is brittle.
3. **January has no per-entity Detailed tabs.** That pack carries `SCI` / `SFP`
   holding the consolidated group statement in a single column. It is reported as
   group-only and skipped rather than silently producing one entity's worth of
   figures. Every other month has all eleven.

Sign convention: figures are extracted exactly as the client states them —
debit-positive, so income and equity are negative. `amount_presentation` applies
`statement_line.sign_multiplier` so it can be compared to our marts directly.
"""
import argparse
import collections
import csv
import datetime
import os
import re
import sys

import openpyxl

# --------------------------------------------------------------------------- paths
SCRIPTS = os.path.dirname(os.path.abspath(__file__))
DATAMODEL = os.path.dirname(SCRIPTS)
ROOT = os.path.dirname(DATAMODEL)
PACKS = os.path.join(ROOT, 'Finance Templates', '2026 TBs', 'Consolidated Accounts')
SEEDS = os.path.join(DATAMODEL, 'seeds')
INTERNAL = os.path.join(ROOT, 'Internal')
OUTDIR = os.path.join(SCRIPTS, 'out')

# --------------------------------------------------------------------------- config
# Client column header -> our company_name. Same vocabulary as
# Scripts/mapping/chain.py; kept here so this script stands alone.
COMPANY = {
    'zaac': 'ZAAC', 'zamre': 'ZAMRE', 'zarib': 'ZARIB',
    'zpal malawi': 'MALAWI', 'zpal': 'MALAWI', 'malawi': 'MALAWI',
    'zamara mena': 'MENA', 'mena': 'MENA',
    'zamara nigeria': 'NIGERIA', 'nigeria': 'NIGERIA',
    'zaaib rwanda': 'RWANDA', 'zamara rwanda': 'RWANDA', 'rwanda': 'RWANDA',
    'zhl': 'ZHL', 'zamara holdings': 'ZHL',
    'zatl': 'ZATL', 'zamara tanzania': 'ZATL',
    'drc': 'DRC', 'zamara drc': 'DRC',
    'c & p': 'C&P', 'c&p': 'C&P', 'c& p': 'C&P', 'c &p': 'C&P', 'c and p': 'C&P',
}
# 'Zamara Holdings Limited' is the pack's own title in A1, never a data column.
TITLE_NOT_ENTITY = {'zamara holdings limited'}

ENTITY_ORDER = ['ZAAC', 'ZAMRE', 'ZARIB', 'MALAWI', 'MENA', 'NIGERIA', 'RWANDA',
                'ZHL', 'ZATL', 'DRC', 'C&P']

# statement -> candidate tab names, best first
DET_TABS = {'SCI': ['SCI Detailed', 'SCI'], 'SFP': ['SFP Detailed', 'SFP']}

MONTHS = {'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
          'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7,
          'aug': 8, 'august': 8, 'sep': 9, 'sept': 9, 'september': 9, 'oct': 10,
          'october': 10, 'nov': 11, 'november': 11, 'dec': 12, 'december': 12}

# Client labels that name one of our statement lines differently, including their
# own typo ('Transalation reserve'). Mirrors Scripts/mapping/derive_map.LABEL_ALIAS.
LABEL_ALIAS = {
    'office cash': 'Cash - Office Cash',
    'client cash': 'Cash - Client Cash',
    'deposits': 'Cash - Deposits',
    'placements': 'Cash - Placements',
    'intangible assets - computer software': 'Intangible - computer software',
    'intangible assets - goodwill': 'Intangible - goodwill',
    'transalation reserve': 'Translation reserve',
    'translation reserve': 'Translation reserve',
    'taxation expense': 'Taxation',
    'finance costs': 'Finance cost',
    # the client's SFP carries the current-year result as an equity line
    'net profit': 'Net Profit (derived)',
}

# Rows that are the client's own arithmetic, not a statement line. They are still
# extracted (row_type='total'/'derived') because they are what the Totals Check
# tab compares against — they just never carry a statement_line_code.
TOTAL_LABELS = {
    'total', 'total income', 'total expenses', 'total assets', 'total equity',
    'total equity and liabilities', 'net assets', 'gross profit', 'operating profit',
    'profit before tax', 'profit after tax', 'profit as per management accounts',
    'variance', 'equity attributable to equity holders of the parent',
    'total non-current assets', 'total current assets', 'total liabilities',
    'total current liabilities', 'total non-current liabilities',
}
# 'Net Profit' on the SFP is =SUM(entire P&L range). Our model derives it in
# fct_trial_balance as net_profit, so we do want to compare it — but it must never
# be treated as a line that owns accounts.
DERIVED_LABELS = {'net profit'}

SECTION_LABELS = {'income', 'expenses', 'assets', 'equity', 'equity and liabilities',
                  'non-current assets', 'current assets', 'non current assets',
                  'current liabilities', 'non-current liabilities', 'owners equity',
                  'liabilities', 'fixed assets'}


def norm(s):
    """Lower-case, collapse whitespace, strip non-breaking spaces."""
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s).replace('\xa0', ' ').strip().lower())


def period_from_name(fname):
    """'March 2026 Consolidated Accounts.xlsx' -> ('2026-03', '2026-03-31')."""
    n = norm(fname)
    ym = re.search(r'(20\d{2})', n)
    if not ym:
        return None, None
    year = int(ym.group(1))
    month = None
    for token in re.findall(r'[a-z]+', n):
        if token in MONTHS:
            month = MONTHS[token]
            break
    if month is None:
        return None, None
    last = (datetime.date(year + (month == 12), month % 12 + 1, 1)
            - datetime.timedelta(days=1))
    return f'{year:04d}-{month:02d}', last.isoformat()


def discover_packs(folder):
    """[(period, period_end, path)] sorted by period. New months need no code change."""
    found = []
    for fn in sorted(os.listdir(folder)):
        if not fn.lower().endswith(('.xlsx', '.xlsm')) or fn.startswith('~$'):
            continue
        period, end = period_from_name(fn)
        if period:
            found.append((period, end, os.path.join(folder, fn)))
    return sorted(found)


def load_statement_lines():
    """(norm line_label -> code, code -> row)."""
    by_label, codes = {}, {}
    path = os.path.join(SEEDS, 'reference', 'statement_line.csv')
    with open(path, newline='', encoding='utf-8-sig') as f:
        for r in csv.DictReader(f):
            by_label[norm(r['line_label'])] = r['statement_line_code']
            codes[r['statement_line_code']] = r
    return by_label, codes


def label_to_code(label, by_label):
    n = norm(label)
    if n in TOTAL_LABELS or n in SECTION_LABELS:
        return None
    return by_label.get(norm(LABEL_ALIAS.get(n, label)))


def entity_columns(rows):
    """Scan the first rows and keep the one with the most entity names.

    Returns ({company: 0-based col}, total_col, header_row_index).
    """
    best, best_total, best_row = {}, None, None
    for hi, row in enumerate(rows[:9]):
        hit, total = {}, None
        for c, v in enumerate(row):
            n = norm(v)
            if not n or n in TITLE_NOT_ENTITY:
                continue
            if c < 2:                      # columns A and B are labels on every tab
                continue
            if n in ('total', 'totals'):
                total = total if total is not None else c
                continue
            co = COMPANY.get(n)
            if co and co not in hit:
                hit[co] = c
        if len(hit) > len(best):
            best, best_total, best_row = hit, total, hi
    return best, best_total, best_row


def num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def read_statement(rows, which, by_label):
    """-> (rows_out, ecols, notes)

    rows_out: [{group, line_label, row_type, statement_line_code, values{co: amt},
                client_total, excel_row}]
    """
    ecols, total_col, hrow = entity_columns(rows)
    notes = []
    if len(ecols) < 3:
        return [], ecols, ['group-only: no per-entity columns found on this tab']

    out, group = [], ''
    for ri, row in enumerate(rows):
        if hrow is not None and ri <= hrow:
            continue
        marker = row[0] if len(row) > 0 else None
        label = row[1] if len(row) > 1 else None
        has_marker = isinstance(marker, str) and marker.strip() != ''
        lab = label.strip() if isinstance(label, str) and label.strip() else ''
        vals = {co: num(row[c]) for co, c in ecols.items() if c < len(row)}
        vals = {co: v for co, v in vals.items() if v is not None}
        ctot = num(row[total_col]) if total_col is not None and total_col < len(row) else None

        nl = norm(lab)
        # 'Net Profit' is tested before the marker: on the SFP it carries the
        # 'Balance Sheet' marker like any other line, but it is =SUM(entire P&L
        # range), so it must never be treated as a line that owns accounts.
        if nl in DERIVED_LABELS:
            row_type = 'derived'
        elif has_marker:
            row_type = 'detail'
        elif nl in TOTAL_LABELS or nl.startswith('total'):
            row_type = 'total'
        elif nl in SECTION_LABELS:
            group = lab
            row_type = 'section'
        elif lab:
            group = lab
            row_type = 'group_header'
        elif vals:
            row_type = 'group_subtotal'
        else:
            continue

        if row_type in ('section', 'group_header') and not vals:
            continue
        if not vals and ctot is None:
            continue

        code = None
        if row_type in ('detail', 'derived'):
            code = label_to_code(lab, by_label) or (label_to_code(group, by_label)
                                                    if group else None)
        out.append({'group': group, 'line_label': lab or f'({row_type})',
                    'row_type': row_type, 'statement_line_code': code or '',
                    'values': vals, 'client_total': ctot, 'excel_row': ri + 1})
    return out, ecols, notes


def extract(packs, by_label, sl, verbose=True):
    long_rows, coverage, unmapped = [], [], collections.Counter()
    # the client's own TOTAL column, per row, for the cross-foot check
    row_totals = {}
    for period, pend, path in packs:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        names = set(wb.sheetnames)
        for which in ('SCI', 'SFP'):
            tab = next((t for t in DET_TABS[which] if t in names), None)
            if tab is None:
                coverage.append({'period': period, 'statement': which, 'tab': '',
                                 'entities': 0, 'lines': 0, 'status': 'tab not found'})
                continue
            rows = [r for r in wb[tab].iter_rows(values_only=True)]
            parsed, ecols, notes = read_statement(rows, which, by_label)
            status = notes[0] if notes else 'ok'
            for r in parsed:
                for co, amt in r['values'].items():
                    code = r['statement_line_code']
                    mult = float(sl[code]['sign_multiplier']) if code in sl else 1.0
                    long_rows.append({
                        'period': period, 'period_end': pend, 'entity': co,
                        'statement': which, 'section_group': r['group'],
                        'line_label': r['line_label'], 'row_type': r['row_type'],
                        'statement_line_code': code,
                        'amount_client_basis': round(amt, 2),
                        'amount_presentation': round(amt * mult, 2),
                        'source_tab': tab, 'source_row': r['excel_row'],
                        'source_file': os.path.basename(path)})
                if r['client_total'] is not None:
                    row_totals[(period, which, r['excel_row'])] = r['client_total']
                if r['row_type'] == 'detail' and not r['statement_line_code']:
                    unmapped[(which, r['group'], r['line_label'])] += 1
            coverage.append({'period': period, 'statement': which, 'tab': tab,
                             'entities': len(ecols),
                             'lines': len({(r['line_label'], r['excel_row']) for r in parsed}),
                             'status': status})
            if verbose:
                print(f'  {period} {which:3s} {tab:14s} entities={len(ecols):2d} '
                      f'rows={len(parsed):3d} {status}', flush=True)
        wb.close()
    return long_rows, coverage, unmapped, row_totals


def totals_check(long_rows, sl):
    """The client's own total rows vs the sum of the detail rows we extracted.

    This is the extract's self-test. A gap means a detail row was missed, read from
    the wrong column, or double-counted — all of which would otherwise show up
    later as a "mapping difference" that isn't one.

    Two facts about the client's layout are built in, both verified across all
    five months and eleven entities:

    * **`Total Expenses` excludes the Taxation line.** Taxation sits inside the
      expense list, above the total, but outside it — the client's total is
      operating expenses plus finance cost. `Management Expense` IS inside it.
    * **`Total Equity and Liabilities` includes the derived `Net Profit` row**,
      which is the current-year result carried to equity.
    """
    cat = {c: r['category_l1'] for c, r in sl.items()}
    detail = collections.defaultdict(float)
    stated = collections.defaultdict(float)
    for r in long_rows:
        k = (r['period'], r['entity'], r['statement'])
        code = r['statement_line_code']
        amt = r['amount_client_basis']
        if r['row_type'] == 'detail':
            if code == 'taxation':
                detail[k + ('TAXATION',)] += amt
            else:
                detail[k + (cat.get(code, 'UNMAPPED'),)] += amt
        elif r['row_type'] == 'derived' and code == 'net_profit':
            detail[k + ('NET PROFIT',)] += amt
        elif r['row_type'] == 'total':
            stated[k + (norm(r['line_label']),)] += amt

    out = []
    keys = {(p, e, s) for p, e, s, _ in list(detail) + list(stated)}
    for period, entity, which in sorted(keys):
        d = {g: v for (p, e, s, g), v in detail.items()
             if (p, e, s) == (period, entity, which)}
        t = {l: v for (p, e, s, l), v in stated.items()
             if (p, e, s) == (period, entity, which)}
        if which == 'SCI':
            pairs = [('Total Income', d.get('INCOME', 0.0), t.get('total income')),
                     ('Total Expenses (excl. taxation)', d.get('EXPENSES', 0.0),
                      t.get('total expenses'))]
        else:
            pairs = [('Total Assets', d.get('ASSETS', 0.0), t.get('total assets')),
                     ('Total Equity and Liabilities',
                      d.get('EQUITY AND LIABILITIES', 0.0) + d.get('NET PROFIT', 0.0),
                      t.get('total equity and liabilities'))]
        for label, ours, theirs in pairs:
            if theirs is None:
                continue
            out.append({'period': period, 'entity': entity, 'statement': which,
                        'check': label, 'sum_of_detail_rows': round(ours, 2),
                        'client_stated_total': round(theirs, 2),
                        'difference': round(ours - theirs, 2),
                        'unmapped_in_this_statement': round(d.get('UNMAPPED', 0.0), 2)})
    return out


def by_line(long_rows, sl):
    """period x entity x statement_line_code, detail + derived rows only."""
    agg = collections.defaultdict(float)
    for r in long_rows:
        if r['row_type'] not in ('detail', 'derived') or not r['statement_line_code']:
            continue
        agg[(r['period'], r['entity'], r['statement'],
             r['statement_line_code'])] += r['amount_client_basis']
    out = []
    for (period, entity, which, code), amt in sorted(agg.items()):
        meta = sl.get(code, {})
        mult = float(meta.get('sign_multiplier', 1))
        out.append({'period': period, 'entity': entity, 'statement': which,
                    'statement_line_code': code, 'line_label': meta.get('line_label', ''),
                    'category_l1': meta.get('category_l1', ''),
                    'line_order': meta.get('line_order', ''),
                    'client_basis_kes': round(amt, 2),
                    'client_presentation_kes': round(amt * mult, 2)})
    return out


# --------------------------------------------------------------------------- output
def write_csvs(outdir, long_rows, line_rows):
    os.makedirs(outdir, exist_ok=True)
    for name, rows in (('client_statements_long', long_rows), ('client_by_line', line_rows)):
        path = os.path.join(outdir, name + '.csv')
        with open(path, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
        print(f'  wrote {path} ({len(rows)} rows)')


def cross_foot(long_rows, row_totals):
    """Sum of the eleven entity columns vs the client's own TOTAL column, per row.

    Exceptions are the client's arithmetic, not ours — their subtotal formula
    missing a row, or a TOTAL that is a different aggregate (the group Profit
    Before Tax rows include the associate). The per-entity columns are what we
    read, so an exception here does not move our figures; it is a data-quality
    item to raise.
    """
    ours = collections.defaultdict(float)
    meta = {}
    for r in long_rows:
        k = (r['period'], r['statement'], r['source_row'])
        ours[k] += r['amount_client_basis']
        meta[k] = (r['section_group'], r['line_label'], r['row_type'])
    out = []
    for k, theirs in sorted(row_totals.items()):
        if k not in ours:
            continue
        d = ours[k] - theirs
        if abs(d) <= 1.0:
            continue
        group, label, rtype = meta[k]
        out.append({'period': k[0], 'statement': k[1], 'excel_row': k[2],
                    'section_group': group, 'line_label': label, 'row_type': rtype,
                    'sum_of_entity_columns': round(ours[k], 2),
                    'client_total_column': round(theirs, 2),
                    'difference': round(d, 2)})
    return out


def write_workbook(path, long_rows, line_rows, coverage, checks, unmapped, packs,
                   row_totals, xfoot):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HDR = PatternFill('solid', fgColor='1F3864')
    SUB = PatternFill('solid', fgColor='D9E1F2')
    WARN = PatternFill('solid', fgColor='FCE4D6')
    WHITE = Font(color='FFFFFF', bold=True, size=10)
    BOLD = Font(bold=True, size=10)
    THIN = Border(bottom=Side('thin', color='BFBFBF'))
    NUM = '#,##0;[Red](#,##0)'

    wb = Workbook()

    def sheet(title):
        ws = wb.create_sheet(title)
        return ws

    def header(ws, cols, row=1):
        for i, c in enumerate(cols, 1):
            cell = ws.cell(row, i, c)
            cell.fill = HDR
            cell.font = WHITE
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
        ws.freeze_panes = ws.cell(row + 1, 1)

    # ---------------------------------------------------------------- Read Me
    ws = wb.active
    ws.title = 'Read Me'
    ws.column_dimensions['A'].width = 118
    lines = [
        ('The client\'s own SCI and SFP, per entity per month', 'title'),
        ('', ''),
        ('What this is', 'h'),
        ('Every figure on the `SCI Detailed` and `SFP Detailed` tabs of the monthly '
         'Consolidated Accounts packs,', ''),
        ('flattened into one table. This is the client\'s statement as they state it — '
         'nothing of ours is in here.', ''),
        ('It is the benchmark the dbt marts are reconciled against.', ''),
        ('', ''),
        ('Sign convention', 'h'),
        ('`amount_client_basis` is exactly as the client shows it: debit-positive, so '
         'income and equity are negative.', ''),
        ('`amount_presentation` applies statement_line.sign_multiplier, which is the '
         'basis our marts hold and', ''),
        ('the basis the recon compares on.', ''),
        ('', ''),
        ('Tabs', 'h'),
        ('SCI Detail / SFP Detail   the client layout — one row per line, one column '
         'per entity, per month.', ''),
        ('By Statement Line         aggregated to our statement_line_code. This is the '
         'comparison grain.', ''),
        ('Totals Check              the client\'s own total rows vs the sum of the '
         'detail rows we extracted.', ''),
        ('Row Cross-Foot            rows where their eleven entity columns do not sum to '
         'their own TOTAL column.', ''),
        ('Unmapped Labels           client detail lines that match no statement_line — '
         'these are excluded from', ''),
        ('                          By Statement Line, so read this tab before trusting '
         'a total.', ''),
        ('Coverage                  which tab was read for each month and how many '
         'entity columns it carried.', ''),
        ('', ''),
        ('January', 'h'),
        ('The January pack has no per-entity Detailed tabs — its `SCI` / `SFP` hold the '
         'consolidated group', ''),
        ('statement in a single column. It is reported as group-only on Coverage and '
         'carries no rows here.', ''),
        ('', ''),
        ('Re-running it next month', 'h'),
        ('Drop the new pack into `Finance Templates/2026 TBs/Consolidated Accounts/` and run:', ''),
        ('    python Scripts/extract_client_statements.py', 'code'),
        ('The month is read from the file name; no code change is needed for a new month.', ''),
        ('', ''),
        (f'Generated {datetime.date.today().isoformat()} from '
         f'{len(packs)} pack(s) by Scripts/extract_client_statements.py', 'small'),
    ]
    r = 1
    for text, kind in lines:
        cell = ws.cell(r, 1, text)
        if kind == 'title':
            cell.font = Font(bold=True, size=14, color='1F3864')
        elif kind == 'h':
            cell.font = Font(bold=True, size=11, color='1F3864')
        elif kind == 'code':
            cell.font = Font(name='Consolas', size=10)
        elif kind == 'small':
            cell.font = Font(italic=True, size=9, color='808080')
        else:
            cell.font = Font(size=10)
        cell.alignment = Alignment(wrap_text=False, vertical='top')
        r += 1

    # ------------------------------------------------- SCI Detail / SFP Detail
    ents = [e for e in ENTITY_ORDER
            if any(r['entity'] == e for r in long_rows)]
    for which in ('SCI', 'SFP'):
        ws = sheet(f'{which} Detail')
        cols = (['Period', 'Section / Group', 'Line', 'Row type', 'Statement line code']
                + ents + ['Row total (ours)', 'Client TOTAL col', 'Diff'])
        header(ws, cols)
        # rebuild the client's row order from source_row
        keyed = collections.defaultdict(dict)
        meta = {}
        for r0 in long_rows:
            if r0['statement'] != which:
                continue
            k = (r0['period'], r0['source_row'])
            keyed[k][r0['entity']] = r0['amount_client_basis']
            meta[k] = (r0['section_group'], r0['line_label'], r0['row_type'],
                       r0['statement_line_code'])
        rr = 2
        for (period, srow) in sorted(keyed):
            group, label, rtype, code = meta[(period, srow)]
            vals = keyed[(period, srow)]
            ws.cell(rr, 1, period)
            ws.cell(rr, 2, group)
            ws.cell(rr, 3, label)
            ws.cell(rr, 4, rtype)
            ws.cell(rr, 5, code)
            tot = 0.0
            for i, e in enumerate(ents):
                v = vals.get(e)
                if v is not None:
                    c = ws.cell(rr, 6 + i, v)
                    c.number_format = NUM
                    tot += v
            c = ws.cell(rr, 6 + len(ents), tot)
            c.number_format = NUM
            ctot = row_totals.get((period, which, srow))
            if ctot is not None:
                c = ws.cell(rr, 7 + len(ents), ctot)
                c.number_format = NUM
                c = ws.cell(rr, 8 + len(ents), round(tot - ctot, 2))
                c.number_format = NUM
                if abs(tot - ctot) > 1:
                    c.fill = WARN
            if rtype in ('total', 'group_subtotal', 'derived'):
                for i in range(1, 6 + len(ents) + 3):
                    ws.cell(rr, i).font = BOLD
                    ws.cell(rr, i).fill = SUB
            rr += 1
        ws.column_dimensions['A'].width = 9
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 34
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 26
        for i in range(len(ents) + 3):
            ws.column_dimensions[get_column_letter(6 + i)].width = 15
        ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}{rr - 1}'

    # ------------------------------------------------------- By Statement Line
    ws = sheet('By Statement Line')
    cols = list(line_rows[0].keys())
    header(ws, [c.replace('_', ' ').title() for c in cols])
    for i, r0 in enumerate(line_rows, 2):
        for j, c in enumerate(cols, 1):
            cell = ws.cell(i, j, r0[c])
            if 'kes' in c:
                cell.number_format = NUM
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = 24 if 'label' in c or 'code' in c else 15
    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}{len(line_rows) + 1}'

    # ------------------------------------------------------------ Totals Check
    ws = sheet('Totals Check')
    cols = list(checks[0].keys()) if checks else ['period']
    header(ws, [c.replace('_', ' ').title() for c in cols])
    for i, r0 in enumerate(checks, 2):
        for j, c in enumerate(cols, 1):
            cell = ws.cell(i, j, r0[c])
            if isinstance(r0[c], float):
                cell.number_format = NUM
        if abs(r0.get('difference', 0)) > 1:
            for j in range(1, len(cols) + 1):
                ws.cell(i, j).fill = WARN
    for j in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 20

    # ---------------------------------------------------------- Row Cross-Foot
    ws = sheet('Row Cross-Foot')
    ws.cell(1, 1, 'Rows where the eleven entity columns do not sum to the client\'s own '
                  'TOTAL column. This is their arithmetic, not ours — we read the entity '
                  'columns, so our figures are unaffected. Raise as a data-quality item.')
    ws.cell(1, 1).font = Font(italic=True, size=9, color='808080')
    xcols = ['period', 'statement', 'excel_row', 'section_group', 'line_label',
             'row_type', 'sum_of_entity_columns', 'client_total_column', 'difference']
    for i, c in enumerate(xcols, 1):
        cell = ws.cell(3, i, c.replace('_', ' ').title())
        cell.fill = HDR
        cell.font = WHITE
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for i, r0 in enumerate(xfoot, 4):
        for j, c in enumerate(xcols, 1):
            cell = ws.cell(i, j, r0[c])
            if c in ('sum_of_entity_columns', 'client_total_column', 'difference'):
                cell.number_format = NUM
        ws.cell(i, len(xcols)).fill = WARN
    for j, w in enumerate([9, 11, 11, 24, 34, 16, 18, 18, 16], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(4, 1)

    # --------------------------------------------------------- Unmapped Labels
    ws = sheet('Unmapped Labels')
    header(ws, ['Statement', 'Section / Group', 'Line label', 'Months seen',
                'Why it matters'])
    for i, ((which, group, label), n) in enumerate(unmapped.most_common(), 2):
        ws.cell(i, 1, which)
        ws.cell(i, 2, group)
        ws.cell(i, 3, label)
        ws.cell(i, 4, n)
        ws.cell(i, 5, 'excluded from By Statement Line and from the recon')
    for j, w in enumerate([10, 26, 40, 12, 46], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---------------------------------------------------------------- Coverage
    ws = sheet('Coverage')
    header(ws, ['Period', 'Statement', 'Tab read', 'Entity columns', 'Line rows', 'Status'])
    for i, r0 in enumerate(coverage, 2):
        ws.cell(i, 1, r0['period'])
        ws.cell(i, 2, r0['statement'])
        ws.cell(i, 3, r0['tab'])
        ws.cell(i, 4, r0['entities'])
        ws.cell(i, 5, r0['lines'])
        ws.cell(i, 6, r0['status'])
        if r0['status'] != 'ok':
            for j in range(1, 7):
                ws.cell(i, j).fill = WARN
    for j, w in enumerate([10, 12, 16, 14, 12, 60], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f'  wrote {path}')


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--packs', default=PACKS, help='folder of Consolidated Accounts packs')
    ap.add_argument('--period', help='extract one period only, e.g. 2026-07')
    ap.add_argument('--xlsx', default=os.path.join(INTERNAL,
                    'Phase1_Client_SCI_SFP_Extract.xlsx'))
    ap.add_argument('--outdir', default=OUTDIR, help='where the CSVs go')
    ap.add_argument('--quiet', action='store_true')
    a = ap.parse_args()

    if not os.path.isdir(a.packs):
        sys.exit(f'pack folder not found: {a.packs}')
    packs = discover_packs(a.packs)
    if a.period:
        packs = [p for p in packs if p[0] == a.period]
    if not packs:
        sys.exit('no packs matched')

    by_label, sl = load_statement_lines()
    print(f'reading {len(packs)} pack(s) from {a.packs}')
    long_rows, coverage, unmapped, row_totals = extract(packs, by_label, sl,
                                                        verbose=not a.quiet)
    if not long_rows:
        sys.exit('nothing extracted — check Coverage: no per-entity Detailed tabs found')
    line_rows = by_line(long_rows, sl)
    checks = totals_check(long_rows, sl)
    xfoot = cross_foot(long_rows, row_totals)

    write_csvs(a.outdir, long_rows, line_rows)
    write_workbook(a.xlsx, long_rows, line_rows, coverage, checks, unmapped, packs,
                   row_totals, xfoot)

    periods = sorted({r['period'] for r in long_rows})
    print()
    print(f'periods extracted   : {", ".join(periods)}')
    print(f'entities            : {len(set(r["entity"] for r in long_rows))}')
    print(f'line cells          : {len(long_rows):,}')
    print(f'statement-line cells: {len(line_rows):,}')
    print(f'unmapped labels     : {len(unmapped)}')
    bad = [c for c in checks if abs(c['difference']) > 1]
    print(f'totals checks off   : {len(bad)} of {len(checks)}')
    print(f'row cross-foot excs : {len(xfoot)} (the client\'s own TOTAL column)')
    for x in xfoot[:6]:
        print('   %-8s %-4s row %-4s %-30s entities %16s vs TOTAL %16s' % (
            x['period'], x['statement'], x['excel_row'], x['line_label'][:30],
            format(x['sum_of_entity_columns'], ',.2f'),
            format(x['client_total_column'], ',.2f')))
    for c in bad[:10]:
        print('   %-8s %-8s %-16s ours %16s vs client %16s' % (
            c['period'], c['entity'], c['check'],
            format(c['sum_of_detail_rows'], ',.2f'),
            format(c['client_stated_total'], ',.2f')))


if __name__ == '__main__':
    main()
