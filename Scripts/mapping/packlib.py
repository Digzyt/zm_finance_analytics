#!/usr/bin/env python3
"""Fast reader for the Zamara monthly Consolidated Accounts packs.

Why this exists: load_tb.py opens workbooks with read_only=True and then uses
ws.cell(r, c) random access, which is O(n^2) on a read-only worksheet (~35s per
pack). Here every sheet is materialised once with iter_rows(values_only=True),
which is ~0.2s per pack. Column/header detection logic mirrors load_tb.py.
"""
import os, re, glob, hashlib, datetime
import openpyxl

# Paths are derived from this file's location so the scripts run in any checkout.
_HERE = os.path.dirname(os.path.abspath(__file__))          # datamodel/Scripts/mapping
SCRIPTS = os.path.dirname(_HERE)
DATAMODEL = os.path.dirname(SCRIPTS)
ROOT = os.path.dirname(DATAMODEL)
SEEDS = os.path.join(DATAMODEL, 'seeds')
OUTDIR = os.path.join(SCRIPTS, 'out')
BASE = os.path.join(ROOT, 'Finance Templates', '2026 TBs', 'Consolidated Accounts')
MONTH_FILES = [
    ('2026-01', '2026-01-31', 'Jan 2026 Consolidated Accounts.xlsx'),
    ('2026-02', '2026-02-28', 'Feb 2026 Consolidated Accounts.xlsx'),
    ('2026-03', '2026-03-31', 'March 2026 Consolidated Accounts.xlsx'),
    ('2026-04', '2026-04-30', 'April 2026 Consolidated Accounts.xlsx'),
    ('2026-05', '2026-05-31', 'May 2026 Consolidated Accounts.xlsx'),
    ('2026-06', '2026-06-30', 'June 2026 Consolidated Accounts.xlsx'),
]

# tab -> (seed name, local currency)
ENTITIES = {
    'ZAAC TB':   ('zaac',    'KES'),
    'ZARIB TB':  ('zarib',   'KES'),
    'Zamre TB':  ('zamre',   'KES'),
    'ZHL':       ('zhl',     'KES'),
    'C & P':     ('c_p',     'KES'),
    'Rwanda':    ('rwanda',  'RWF'),
    'Nigeria':   ('nigeria', 'NGN'),
    'Malawi TB': ('malawi',  'MWK'),
    'MENA TB':   ('mena',    'AED'),
    'DRC TB':    ('drc',     'USD'),
    'ZATL':      ('zatl',    'TZS'),
}

CODE_HDRS = {'a/c no', 'bc code', 'bc codes', 'account no', 'code'}
DESC_HDRS = {'description', 'account name', 'account'}
AMT_POST  = {'amount after accruals', 'amount after accrual', 'amount after accural',
             'net amount',
             # ZARIB, March and April: the client renames both columns
             'amount after deferred rev', 'amount after deffered rev',
             'amount after deferred revenue', 'amount after def rev', 'amount after accruals '}
AMT_PLAIN = {'amount', 'dr', 'tzs', 'net debit /(credit)', 'net debit/(credit)', 'local', 'debit'}
CR_HDRS   = {'cr', 'credit'}
KES_HDRS  = {'kes'}
# See reseed_from_packs.ACCRUAL_HDRS: where the client's netted column has no
# header we must net Amount + Accruals ourselves, or accrual-only rows (empty
# Amount) are dropped.
ACCRUAL_HDRS = {'accrual', 'accruals',
                # ZARIB March/April; the misspelling is the client's
                'deffered revenue & accruals', 'deferred revenue & accruals',
                'deffered revenue and accruals', 'deferred revenue and accruals',
                'deferred revenue & accrual'}

SKIP_DESC = {'assets', 'non-current assets', 'current assets', 'fixed assets', 'liabilities',
    'equity', 'equity and liabilities', 'owners equity', 'current liabilities',
    'non-current liabilities', 'income', 'expenses', 'total', 'total assets', 'total equity',
    'total equity and liabilities', 'total income', 'total expenses', ''}


def norm(s):
    return re.sub(r'\s+', ' ', str(s).replace('\xa0', ' ').strip().lower()) if s is not None else ''


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return None
    try:
        return float(str(v).replace(',', '').strip())
    except Exception:
        return None


def sheet_rows(path, wanted=None):
    """Return {sheet_name: [row_tuples]} materialised once."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for name in wb.sheetnames:
        if wanted is not None and name not in wanted:
            continue
        out[name] = [r for r in wb[name].iter_rows(values_only=True)]
    wb.close()
    return out


def find_header(rows, maxscan=12):
    """(0-based header row index, {role: 0-based col index}) mirroring load_tb.find_header."""
    for i, row in enumerate(rows[:maxscan]):
        vals = {c: norm(v) for c, v in enumerate(row[:14])}
        if not any(v in DESC_HDRS for v in vals.values()):
            continue
        roles = {}
        for c, v in vals.items():
            if v in CODE_HDRS and 'code' not in roles: roles['code'] = c
            elif v in DESC_HDRS and 'desc' not in roles: roles['desc'] = c
            elif v in KES_HDRS and 'kes' not in roles: roles['kes'] = c
            elif v in AMT_POST and 'amt_post' not in roles: roles['amt_post'] = c
            elif v in CR_HDRS and 'cr' not in roles: roles['cr'] = c
            elif v in ACCRUAL_HDRS and 'accrual' not in roles: roles['accrual'] = c
            elif v in AMT_PLAIN and 'amt_plain' not in roles: roles['amt_plain'] = c
        if 'desc' in roles and ('amt_post' in roles or 'amt_plain' in roles):
            return i, roles
    return None, None


def read_entity(rows):
    """-> (roles, [(code_or_None, description, local_amount, kes_amount)])"""
    hi, roles = find_header(rows)
    if hi is None:
        return None, None
    ccode = roles.get('code'); cdesc = roles['desc']
    camt = roles.get('amt_post') if 'amt_post' in roles else roles.get('amt_plain')
    ccr = roles.get('cr'); ckes = roles.get('kes')
    # only net it ourselves when the client has not supplied a netted column
    cacc = roles.get('accrual') if 'amt_post' not in roles else None
    out = []
    for row in rows[hi + 1:]:
        def cell(c):
            return row[c] if c is not None and c < len(row) else None
        desc = cell(cdesc)
        nd = norm(desc)
        if nd in SKIP_DESC or nd.startswith('total') or nd.startswith('period'):
            continue
        loc = num(cell(camt))
        acc = num(cell(cacc)) if cacc is not None else None
        if acc is not None:
            loc = (loc or 0.0) + acc
        if ccr is not None:
            loc = (loc or 0.0) - (num(cell(ccr)) or 0.0)
        kes = num(cell(ckes)) if ckes is not None else None
        if loc is None and kes is None:
            continue
        code = None
        if ccode is not None and cell(ccode) not in (None, ''):
            code = str(cell(ccode)).strip()
        out.append((code, str(desc).strip() if desc else '', loc or 0.0, kes))
    return roles, out


RATE_CURR = {'rwanda franc': 'RWF', 'malawi kwacha': 'MWK', 'ugx shs': 'UGX',
             'nigeria naira': 'NGN', 'tanzania shs': 'TZS', 'usd': 'USD', 'aed': 'AED'}


def read_rates(rows, period):
    if rows is None:
        return None
    hdr = None; cols = {}
    for i, row in enumerate(rows[:12]):
        vals = {c: norm(v) for c, v in enumerate(row)}
        hit = {c: RATE_CURR[v] for c, v in vals.items() if v in RATE_CURR}
        if len(hit) >= 3:
            hdr = i; cols = hit; break
    if hdr is None:
        return None
    closing = None
    for row in rows[hdr + 1:]:
        d = row[0] if row else None
        if isinstance(d, datetime.datetime) and f'{d.year:04d}-{d.month:02d}' == period:
            closing = row
    if closing is None:
        return None
    out = {'KES': 1.0}
    for c, cur in cols.items():
        v = num(closing[c]) if c < len(closing) else None
        if v:
            out[cur] = v
    return out


def synth_code(desc):
    return 'X-' + hashlib.md5(str(desc).encode('utf-8')).hexdigest()[:10].upper()
