#!/usr/bin/env python3
"""Resolve the client's own formula chain in a Consolidated Accounts pack.

    SCI/SFP Detailed  ->  KES consolidated TB  ->  TB Local Currency  ->  entity TB tab

Every figure the client reports is a formula that terminates in a specific cell
on an entity's TB tab, so the account -> statement-line assignment is theirs,
not something we infer from account names.

Three things this has to get right, all of which bit earlier attempts:

* `KES consolidated TB` and `TB Local Currency` are NOT reliably row-aligned,
  and the offset differs per entity (C&P is commonly a couple of rows out). The
  row link is therefore taken from the formula itself — including resolving
  VLOOKUPs by their lookup key — never by matching account names.
* `TB Local Currency` covers only 9 of the 11 entities. ZHL and ZATL have no
  column there; their `KES consolidated TB` cells point straight at `ZHL!C39`,
  `ZATL!D6` and so on.
* The entity column order differs between tabs (SCI Detailed starts at column C,
  SFP Detailed and KES consolidated TB at column D), so every tab's own header
  row is read rather than assuming a layout.
"""
import re, collections
import openpyxl
import packlib as P

SRC_TABS = {'ZAAC TB', 'ZARIB TB', 'Zamre TB', 'ZHL', 'C & P', 'Rwanda', 'Nigeria',
            'Malawi TB', 'MENA TB', 'DRC TB', 'ZATL'}
KES_TAB = 'KES consolidated TB'
LOC_TAB = 'TB Local Currency'
DET = {'SCI': ['SCI Detailed', 'SCI'], 'SFP': ['SFP Detailed', 'SFP']}

COMPANY = {
    'zaac': 'ZAAC', 'zamre': 'ZAMRE', 'zarib': 'ZARIB',
    'zpal malawi': 'MALAWI', 'zpal': 'MALAWI', 'malawi': 'MALAWI',
    'zamara mena': 'MENA', 'mena': 'MENA',
    'zamara nigeria': 'NIGERIA', 'nigeria': 'NIGERIA',
    'zaaib rwanda': 'RWANDA', 'zamara rwanda': 'RWANDA', 'rwanda': 'RWANDA',
    'zhl': 'ZHL', 'zamara holdings': 'ZHL', 'zamara holdings limited': 'ZHL',
    'zatl': 'ZATL', 'zamara tanzania': 'ZATL',
    'drc': 'DRC', 'zamara drc': 'DRC',
    'c & p': 'C&P', 'c&p': 'C&P', 'c& p': 'C&P', 'c &p': 'C&P', 'c and p': 'C&P',
}
TAB2CO = {'ZAAC TB': 'ZAAC', 'ZARIB TB': 'ZARIB', 'Zamre TB': 'ZAMRE', 'ZHL': 'ZHL',
          'C & P': 'C&P', 'Rwanda': 'RWANDA', 'Nigeria': 'NIGERIA', 'Malawi TB': 'MALAWI',
          'MENA TB': 'MENA', 'DRC TB': 'DRC', 'ZATL': 'ZATL'}

XCELL = re.compile(r"(?:'([^']+)'|([A-Za-z][A-Za-z0-9 &_.]*))!\$?([A-Z]{1,3})\$?(\d+)(?!\s*:)")
XRANGE = re.compile(r"(?:'([^']+)'|([A-Za-z][A-Za-z0-9 &_.]*))!\$?([A-Z]{1,3})\$?\d*:\$?([A-Z]{1,3})\$?\d*")
LOCAL_CELL = re.compile(r"(?<![!:$A-Z0-9])\$?([A-Z]{1,2})\$?(\d+)(?![(!:0-9])")
VLOOKUP = re.compile(r"VLOOKUP\s*\(([^,]+),([^,]+),\s*(\d+)", re.I)


def cidx(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


XRANGE_FULL = re.compile(
    r"(?:'([^']+)'|([A-Za-z][A-Za-z0-9 &_.]*))!"
    r"\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)")

MAX_RANGE = 130   # a SUM over more rows than this is a whole-statement subtotal


def ext_refs(formula, expand_ranges=True):
    """External single-cell refs as [(sheet, col_idx, row)].

    Ranges have to be expanded, not dropped: the client writes much of the SFP as
    SUM('KES consolidated TB'!D14:D28), and dropping those loses most of the
    balance-sheet mapping. Only single-column ranges are expanded — a
    multi-column range is a VLOOKUP table, not a list of accounts — and anything
    longer than MAX_RANGE rows is treated as a subtotal over a whole statement
    (SFP 'Net Profit' sums the entire P&L) and ignored.
    """
    if not isinstance(formula, str) or not formula.startswith('='):
        return []
    out = []
    if expand_ranges:
        for m in XRANGE_FULL.finditer(formula):
            sheet = (m.group(1) or m.group(2)).strip()
            c1, r1, c2, r2 = cidx(m.group(3)), int(m.group(4)), cidx(m.group(5)), int(m.group(6))
            if c1 != c2 or abs(r2 - r1) + 1 > MAX_RANGE:
                continue
            for r in range(min(r1, r2), max(r1, r2) + 1):
                out.append((sheet, c1, r))
    masked = XRANGE.sub(lambda m: ' ' * len(m.group(0)), formula)
    out += [((m.group(1) or m.group(2)).strip(), cidx(m.group(3)), int(m.group(4)))
            for m in XCELL.finditer(masked)]
    return out


class Pack:
    def __init__(self, path, period):
        self.period = period
        self.path = path
        self.wbF = openpyxl.load_workbook(path, data_only=False)
        self.wbV = openpyxl.load_workbook(path, data_only=True)
        self.names = set(self.wbF.sheetnames)
        self.ecols = {t: self._entity_cols(t) for t in
                      [KES_TAB, LOC_TAB] + [x for v in DET.values() for x in v]
                      if t in self.names}
        self._loc_key_index = {}
        self._src = self._source_index()

    def close(self):
        self.wbF.close(); self.wbV.close()

    def f(self, tab, r, c):
        return self.wbF[tab].cell(r, c).value

    def v(self, tab, r, c):
        return self.wbV[tab].cell(r, c).value

    def _entity_cols(self, tab):
        """Pick the single header row with the most entity names in it.

        Scanning every row and taking first-seen wins picks up the sheet title —
        'Zamara Holdings Limited' in A1 was being read as ZHL's data column,
        which silently emptied ZHL's whole chain. Columns A and B are labels on
        every one of these tabs, so they are never data columns.
        """
        ws = self.wbV[tab]
        best = {}
        for hr in range(1, 9):
            row = {}
            for c in range(3, min(ws.max_column, 24) + 1):
                co = COMPANY.get(P.norm(ws.cell(hr, c).value))
                if co and co not in row:
                    row[co] = c
            if len(row) > len(best):
                best = row
        return best

    # ------------------------------------------------------- source tab index
    def _source_index(self):
        """(tab, row) -> (source_code, description), using the same header and
        column detection as the reseed so the keys match the bronze seeds."""
        idx = {}
        for tab in SRC_TABS & self.names:
            ws = self.wbV[tab]
            rows = [tuple(ws.cell(r, c).value for c in range(1, ws.max_column + 1))
                    for r in range(1, ws.max_row + 1)]
            hi, roles = P.find_header(rows)
            if hi is None:
                continue
            ccode = roles.get('code'); cdesc = roles['desc']
            for i, row in enumerate(rows):
                if i <= hi:
                    continue
                desc = row[cdesc] if cdesc < len(row) else None
                if desc is None or P.norm(desc) in P.SKIP_DESC:
                    continue
                code = None
                if ccode is not None and ccode < len(row) and row[ccode] not in (None, ''):
                    code = str(row[ccode]).strip()
                idx[(tab, i + 1)] = (code, str(desc).strip())
        return idx

    # -------------------------------------------- TB Local Currency resolution
    def _loc_lookup(self, key_col, value):
        """Row on TB Local Currency whose key_col equals value (VLOOKUP semantics)."""
        if key_col not in self._loc_key_index:
            ws = self.wbV[LOC_TAB]
            m = {}
            for r in range(1, ws.max_row + 1):
                k = P.norm(ws.cell(r, key_col).value)
                if k and k not in m:
                    m[k] = r
            self._loc_key_index[key_col] = m
        return self._loc_key_index[key_col].get(P.norm(value))

    def loc_accounts(self, loc_row, company):
        """Source accounts a TB Local Currency row resolves to for one entity."""
        col = self.ecols.get(LOC_TAB, {}).get(company)
        if not col or LOC_TAB not in self.names:
            return []
        out = []
        for sheet, c, r in ext_refs(self.f(LOC_TAB, loc_row, col)):
            if sheet in SRC_TABS and TAB2CO.get(sheet) == company:
                out.append((sheet, r))
        return out

    def kes_accounts(self, kes_row, company):
        """Source accounts a KES consolidated TB row resolves to for one entity.
        Handles the three forms the client uses: a direct TB Local Currency cell
        ref, a VLOOKUP into TB Local Currency, and (ZHL / ZATL) a direct ref to
        the entity tab."""
        col = self.ecols.get(KES_TAB, {}).get(company)
        if not col:
            return []
        formula = self.f(KES_TAB, kes_row, col)
        out = []
        for sheet, c, r in ext_refs(formula):
            if sheet in SRC_TABS and TAB2CO.get(sheet) == company:
                out.append((sheet, r))
            elif sheet == LOC_TAB:
                out += self.loc_accounts(r, company)
        if isinstance(formula, str):
            for m in VLOOKUP.finditer(formula):
                key_expr, table, ret = m.group(1), m.group(2), int(m.group(3))
                rng = XRANGE.search(table)
                if not rng or (rng.group(1) or rng.group(2)).strip() != LOC_TAB:
                    continue
                start = cidx(rng.group(3))
                lm = LOCAL_CELL.search(key_expr)
                if not lm:
                    continue
                key_val = self.v(KES_TAB, int(lm.group(2)), cidx(lm.group(1)))
                lr = self._loc_lookup(start, key_val)
                if lr:
                    out += self.loc_accounts(lr, company)
        return out

    # ---------------------------------------------------------------- the chain
    def statement_lines(self, which):
        """[(row, own_label, group_label)] for the SCI/SFP Detailed tab.

        On the SCI the expense section is a group header row (column A blank,
        e.g. 'Personnel Costs') followed by account-level detail rows carrying
        'P&L' in column A. The detail row's own label is an account name, not a
        statement line, so it has to inherit its group. On the SFP every line row
        carries 'Balance' in column A and its own label IS the statement line.
        """
        tab = next((t for t in DET[which] if t in self.names), None)
        if tab is None:
            return None, [], {}
        ws = self.wbV[tab]
        lines = []
        group = ''
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            lab = ws.cell(r, 2).value
            if not (isinstance(lab, str) and lab.strip()) or P.norm(lab) in P.SKIP_DESC:
                continue
            lab = lab.strip()
            # Column A is the client's marker column: 'P&L' on the SCI,
            # 'Balance Sheet' on the SFP. Group headers and subtotals leave it
            # blank. Testing for a non-empty string is the structural rule —
            # matching literal marker text is brittle (the SFP marker is
            # 'Balance Sheet', not 'Balance').
            is_detail = isinstance(a, str) and a.strip() != ''
            if not is_detail:
                group = lab
                continue
            lines.append((r, lab, group))
        return tab, lines, self.ecols.get(tab, {})

    def build(self):
        """-> line_accounts {(company, SCI|SFP, line_label): [(src_code, norm_desc)]},
              categories  {(company, src_code, norm_desc): client Categories label},
              reported    {(company, SCI|SFP, line_label): value}"""
        line_accounts = collections.defaultdict(list)
        reported = {}
        # MENA has no account codes; stg_mena_descriptive_tb keys it on
        # 'MENA-' || md5(Description), so the raw description string has to be
        # carried through byte-for-byte, not just its normalised form.
        raw_desc = {}
        for which in ('SCI', 'SFP'):
            tab, lines, ecols = self.statement_lines(which)
            if tab is None:
                continue
            for r, label, group in lines:
                for co, c in ecols.items():
                    val = self.v(tab, r, c)
                    if isinstance(val, (int, float)):
                        # keyed on BOTH labels so the caller can pick exactly one
                        # per row — counting a row under its own label and again
                        # under its group double-counts (the SCI has a 'Taxation'
                        # group over a 'Taxation expense' row, both of which name
                        # the same statement line)
                        k = (co, which, label, group)
                        reported[k] = reported.get(k, 0.0) + float(val)
                    seen = set()
                    for sheet, rc, rr in ext_refs(self.f(tab, r, c)):
                        targets = []
                        if sheet == KES_TAB:
                            targets = self.kes_accounts(rr, co)
                        elif sheet == LOC_TAB:
                            targets = self.loc_accounts(rr, co)
                        elif sheet in SRC_TABS and TAB2CO.get(sheet) == co:
                            targets = [(sheet, rr)]
                        for t, rw in targets:
                            acc = self._src.get((t, rw))
                            if not acc:
                                continue
                            key = (acc[0] or '', P.norm(acc[1]))
                            raw_desc[(co,) + key] = acc[1]
                            if key not in seen:
                                seen.add(key)
                                line_accounts[(co, which, label)].append(key)
                                if group:
                                    line_accounts[(co, which, group)].append(key)

        cats = {}
        ws = self.wbV[KES_TAB]
        for r in range(1, ws.max_row + 1):
            cat = ws.cell(r, 1).value
            if not isinstance(cat, str) or not cat.strip():
                continue
            for co in self.ecols.get(KES_TAB, {}):
                for t, rw in self.kes_accounts(r, co):
                    acc = self._src.get((t, rw))
                    if acc:
                        cats.setdefault((co, acc[0] or '', P.norm(acc[1])), cat.strip())
                        raw_desc.setdefault((co, acc[0] or '', P.norm(acc[1])), acc[1])
        return line_accounts, cats, reported, raw_desc


def load(path, period):
    return Pack(path, period)
