#!/usr/bin/env python3
"""Extract the client's debtor analysis into the debtor_analysis seed.

    python Scripts/extract_debtor_analysis.py           # dry run — prints summary
    python Scripts/extract_debtor_analysis.py --write   # writes seeds/reference/debtor_analysis.csv
    dbt seed --full-refresh && dbt build

Source
------
`Internal/Debtors_Master_Data_Template.xlsx`, sheet `Debtors_Data_Entry`. That
workbook is the standardized intake for the client's aged debtor analysis: one
row per client per reporting date, carrying the entity, department, division,
client name and ID, currency, premium in/outstanding (BC), twelve ageing buckets
(0-30 … >3 years), a total, collected-to-date and outstanding-balance columns.
The template is itself fed from `Finance Templates/ZAMARA CONSOLIDATED GROUP
Debtors … .xlsx` by whoever prepares the monthly pack. This script does not
re-derive the analysis from the many heterogeneous entity sheets of that source —
it takes the standardized template as authoritative, and flags anything in it
that does not parse cleanly.

Why a seed at all
-----------------
Debtors arrive by hand (BC is still in progress), so the analysis is a designed
manual input, not something read off a system. Like every other manual input it
travels as a reference seed and is transformed through dbt to the reporting marts
(`rpt_debtor_analysis` / `rpt_debtor_ageing`), which Power BI then consumes. The
seed holds the figures in each entity's own FUNCTIONAL currency, exactly as the
template captures them; nobody translates the aged balances to the group currency
until the FX basis for debtor balances is agreed with Finance.

Conventions
-----------
- Entity names are mapped to the canonical codes in `entity.csv` (Tanzania ->
  ZATL, Malawi -> MALAWI, …) so the marts can join `dim_entity` and `fx_rate`
  cleanly.
- `period` is derived from the reporting date (`01/07/2026` -> `2026-07`) and
  `reporting_date` is stored ISO (`2026-07-01`).
- Numeric columns are coerced; anything that is not a number becomes NULL and is
  reported, because a stray label in a numeric column (see below) must not be
  silently read as an amount.
"""
import argparse
import collections
import csv
import datetime
import os
import re
import sys

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
DATAMODEL = os.path.dirname(_HERE)
ROOT = os.path.dirname(DATAMODEL)
SEED = os.path.join(DATAMODEL, 'seeds', 'reference', 'debtor_analysis.csv')
SRC = os.path.join(ROOT, 'Internal', 'Debtors_Master_Data_Template.xlsx')
SHEET = 'Debtors_Data_Entry'

# template entity label -> canonical entity_code (entity.csv)
ENTITY_MAP = {
    'ZARIB': 'ZARIB',
    'ZAAC': 'ZAAC',
    'ZAMRE': 'ZAMRE',
    'ZHL': 'ZHL',
    'C&P': 'C&P',
    'Malawi': 'MALAWI',
    'DRC': 'DRC',
    'Nigeria': 'NIGERIA',
    'Rwanda': 'RWANDA',
    'Uganda': 'UGANDA',
    'Tanzania': 'ZATL',
    'MENA': 'MENA',
}

# template header -> seed column, in template column order
HEADERS = [
    ('Reporting_Date',      'reporting_date'),
    ('Entity',              'entity'),
    ('Department_Unit',     'department_unit'),
    ('Division Name',       'division_name'),
    ('Client Name',         'client_name'),
    ('Client ID',           'client_id'),
    ('Currency',            'currency'),
    ('Premium Tran BC',     'premium_tran_bc'),
    ('Premium OS BC',       'premium_os_bc'),
    ('0-30 Days',           'age_0_30'),
    ('31-60 Days',          'age_31_60'),
    ('61-90 Days',          'age_61_90'),
    ('91-120 Days',         'age_91_120'),
    ('121-180 Days',        'age_121_180'),
    ('181-240 Days',        'age_181_240'),
    ('241-300 Days',        'age_241_300'),
    ('301-365 Days',        'age_301_365'),
    ('1-2 Years',           'age_1_2y'),
    ('2-3 Years',           'age_2_3y'),
    ('>3 Years',            'age_gt_3y'),
    ('Total',               'total'),
    ('Collections 1st-10th', 'collections_1st_10th'),
    ('DP 1st-10th',         'dp_1st_10th'),
    ('O/S Balance',         'os_balance'),
]

HDR_TO_SEED = {h: s for h, s in HEADERS}

# the twelve ageing bucket headers (numeric)
BUCKET_HDRS = [h for h, s in HEADERS if h.startswith(('0-30', '31-60', '61-90',
                '91-120', '121-180', '181-240', '241-300', '301-365',
                '1-2', '2-3', '>3'))]

FIELDS = (['period', 'company_name', 'reporting_date', 'department_unit',
           'division_name', 'client_name', 'client_id', 'currency',
           'premium_tran_bc', 'premium_os_bc']
          + [HDR_TO_SEED[h] for h in BUCKET_HDRS]
          + ['total', 'collections_1st_10th', 'dp_1st_10th', 'os_balance'])

_NUMERIC_HDRS = (['Premium Tran BC', 'Premium OS BC', 'Total', 'O/S Balance']
                 + BUCKET_HDRS + ['Collections 1st-10th', 'DP 1st-10th'])


def _num(v):
    """-> float or None. Accepts int/float and numeric strings ('1,234.5')."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 4)
    s = str(v).strip()
    if not s or s.lower() in ('null', 'nan', '-'):
        return None
    s = s.replace(',', '')
    try:
        return round(float(s), 4)
    except ValueError:
        return None  # non-numeric cell; caller reports it


def parse_reporting_date(raw):
    """'01/07/2026' -> (period '2026-07', reporting_date '2026-07-01')."""
    if isinstance(raw, datetime.datetime):
        return raw
    s = str(raw).strip()
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f'cannot parse reporting date: {raw!r}')


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--source', default=SRC, help='path to the template workbook')
    ap.add_argument('--seed', default=SEED)
    ap.add_argument('--write', action='store_true')
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.source, read_only=True, data_only=True)
    ws = wb[SHEET]
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else '' for c in next(it)]
    col_of = {h: i for i, h in enumerate(header)}

    missing = [h for h, _ in HEADERS if h not in col_of]
    if missing:
        sys.exit(f'template is missing headers: {missing} — check {SHEET}')

    # ---- reporting date / period -------------------------------------------
    # Scan for the reporting date first (peek at the first non-empty cell); the
    # processing pass below re-reads the whole sheet from the top so no data
    # row is consumed by the scan.
    raw_date = None
    wb2 = openpyxl.load_workbook(a.source, read_only=True, data_only=True)
    ws2 = wb2[SHEET]
    it2 = ws2.iter_rows(values_only=True)
    next(it2)  # skip header
    for row in it2:
        d = row[col_of['Reporting_Date']]
        if d is not None:
            raw_date = d
            break
    wb2.close()
    dt = parse_reporting_date(raw_date)
    period = dt.strftime('%Y-%m')
    reporting_date = dt.strftime('%Y-%m-%d')

    rows = []
    dq = collections.Counter()
    for row in it:
        entity_raw = row[col_of['Entity']]
        client = row[col_of['Client Name']]
        if entity_raw is None and client is None:
            continue
        entity = str(entity_raw).strip() if entity_raw is not None else None
        if entity not in ENTITY_MAP:
            sys.exit(f'unknown entity label {entity!r} — add to ENTITY_MAP')
        client_name = _text(client)
        if client_name is None:
            continue
        r = {'period': period,
             'company_name': ENTITY_MAP[entity],
             'reporting_date': reporting_date,
             'department_unit': _text(row[col_of['Department_Unit']]),
             'division_name': _text(row[col_of['Division Name']]),
             'client_name': client_name,
             'client_id': _text(row[col_of['Client ID']]),
             'currency': _text(row[col_of['Currency']])}
        for h in BUCKET_HDRS + ['Premium Tran BC', 'Premium OS BC', 'Total',
                                'Collections 1st-10th', 'DP 1st-10th',
                                'O/S Balance']:
            v = row[col_of[h]]
            if isinstance(v, str) and v.strip() and _num(v) is None:
                dq[f'{HDR_TO_SEED[h]} <- {v.strip()!r}'] += 1
            r[HDR_TO_SEED[h]] = _num(v)
        rows.append(r)
    wb.close()

    n = len(rows)
    if n == 0:
        sys.exit('no debtor rows read — check the template path and sheet name')

    # ---- verification --------------------------------------------------------
    by_ent = collections.Counter(r['company_name'] for r in rows)
    print(f'read {n} client rows for {period} (reporting {reporting_date})')
    print('entities: ' + ' '.join(f'{e}={by_ent[e]}' for e in sorted(by_ent)))

    print('\nentity totals (stated Total column, functional currency):')
    tot = collections.defaultdict(float)
    for r in rows:
        tot[r['company_name']] += (r['total'] or 0.0)
    for e in sorted(tot):
        print(f'  {e:7s} {tot[e]:>18,.2f}')

    mism = sum(1 for r in rows
               if r['total'] is not None
               and abs(sum(r[HDR_TO_SEED[h]] or 0.0 for h in BUCKET_HDRS)
                       - r['total']) > 0.01)
    print(f'\nrows where the ageing buckets do not sum to the stated Total: {mism}')
    print('  (expected for Uganda/MENA rows whose source Total exceeds the bucket '
          'sum; the stated Total is preserved in the seed)')

    if dq:
        print('\ndata-quality notes (non-numeric cells read as NULL):')
        for k, v in dq.most_common():
            print(f'  {k}: {v}')

    if not a.write:
        print('\ndry run — pass --write to update the seed')
        return

    os.makedirs(os.path.dirname(a.seed), exist_ok=True)
    with open(a.seed, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=FIELDS, extrasaction='ignore')
        w.writeheader()
        w.writerows(rows)
    print(f'\nwrote {a.seed} ({len(rows)} rows)')
    print('now run: dbt seed --full-refresh && dbt build')


def _text(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() in ('null', 'nan'):
        return None
    return s


if __name__ == '__main__':
    main()
