#!/usr/bin/env python3
"""
load_tb.py — add a monthly Trial Balance pack to the dbt bronze seeds.

Point it at ONE workbook (either format works):
  • a "Consolidated Accounts" / "Consolidated TB" workbook  (has statement tabs + per-entity TB tabs)
  • a "Raw TBs" workbook                                     (per-entity TB tabs only)

What it does
------------
1. Parses the reporting PERIOD from the file name (e.g. "June 2026" -> 2026-06-30).
2. For each entity TB tab it finds, auto-detects the header row and columns
   (works whether the tab uses BC codes, legacy local codes, a post-accrual
   column, or is description-only).
3. Reads that entity's year-to-date (YTD) balance per account and computes the
   MONTH's MOVEMENT = new YTD - prior cumulative already in the seed. Bronze
   seeds hold monthly movements (that is what the dbt period model expects).
4. Appends the movement rows (dated to the period month-end) to
   seeds/bronze/gl_entry_<entity>.csv. Re-running for a period that is already
   loaded first removes that period's rows, so it is safe/idempotent.
5. Best-effort: appends the period's FX rates from the "Rates" tab to
   seeds/reference/fx_rate.csv (verify these).

Usage
-----
    python Scripts/load_tb.py "<path to workbook>"            # dry run (prints a summary)
    python Scripts/load_tb.py "<path to workbook>" --write    # write to the seeds

After --write:   dbt seed --full-refresh && dbt build
"""
import sys, os, re, csv, hashlib, argparse, datetime
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SEEDS = os.path.normpath(os.path.join(HERE, '..', 'seeds'))
BRONZE = os.path.join(SEEDS, 'bronze')
REF = os.path.join(SEEDS, 'reference')

MONTHS = {m.lower(): i for i, m in enumerate(
    ['January','February','March','April','May','June','July','August','September','October','November','December'], 1)}
MONTHS.update({'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12})

# Per-entity config. tab -> (seed name, local currency, descriptive?)
# descriptive = the tab has no account-code column; the description is the key.
ENTITIES = {
    'ZAAC TB':  ('zaac',   'KES', False),
    'ZARIB TB': ('zarib',  'KES', False),
    'Zamre TB': ('zamre',  'KES', False),
    'ZHL':      ('zhl',    'KES', False),
    'C & P':    ('c_p',    'KES', False),
    'Rwanda':   ('rwanda', 'RWF', False),
    'Nigeria':  ('nigeria','NGN', False),
    'Malawi TB':('malawi', 'MWK', False),
    'MENA TB':  ('mena',   'AED', True),
    'DRC TB':   ('drc',    'USD', False),
    'ZATL':     ('zatl',   'TZS', False),
}
STD_HEADER = ['Entry_No','Posting_Date','Document_Type','Document_No','External_Document_No','Transaction_No',
    'G_L_Account_No','Description','Amount','Debit_Amount','Credit_Amount','Source_Code','Source_Type','Source_No',
    'Dimension_Set_ID','Global_Dimension_1_Code','Global_Dimension_2_Code','Business_Unit_Code','IC_Partner_Code',
    'User_ID','Reversed','Reversed_by_Entry_No','Reversed_Entry_No']

CODE_HDRS = {'a/c no','bc code','bc codes','account no','code'}
DESC_HDRS = {'description','account name','account'}
# local-amount header preference (post-accrual first), then plain; KES is separate
AMT_POST  = {'amount after accruals','amount after accrual','net amount'}
AMT_PLAIN = {'amount','dr','tzs','net debit /(credit)','net debit/(credit)','local'}
KES_HDRS  = {'kes'}

def norm(s):
    return re.sub(r'\s+', ' ', str(s).replace('\xa0', ' ').strip().lower()) if s is not None else ''

def num(v):
    if isinstance(v, (int, float)): return float(v)
    if v is None: return None
    try: return float(str(v).replace(',', '').strip())
    except Exception: return None

def parse_period(fname):
    base = os.path.basename(fname)
    mm = None; yy = None
    m = re.search(r'([A-Za-z]{3,9})\s*[\'\-\s]*?(20\d{2})', base)
    if m and m.group(1).lower() in MONTHS:
        mm = MONTHS[m.group(1).lower()]; yy = int(m.group(2))
    if mm is None:
        m2 = re.search(r'(20\d{2})[-_/](\d{1,2})', base)
        if m2: yy = int(m2.group(1)); mm = int(m2.group(2))
    if mm is None:
        raise SystemExit(f"Could not parse a month/year from file name: {base!r}. "
                         f"Rename it to include e.g. 'June 2026'.")
    last = [31,29 if yy%4==0 and (yy%100!=0 or yy%400==0) else 28,31,30,31,30,31,31,30,31,30,31][mm-1]
    return f'{yy:04d}-{mm:02d}', f'{yy:04d}-{mm:02d}-{last:02d}'

def find_header(ws, maxscan=12):
    """Return (row_index, {role: col_index}) for the first plausible header row."""
    for r in range(1, min(maxscan, ws.max_row) + 1):
        vals = {c: norm(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 14) + 1)}
        has_desc = any(v in DESC_HDRS for v in vals.values())
        if not has_desc:
            continue
        roles = {}
        for c, v in vals.items():
            if v in CODE_HDRS and 'code' not in roles: roles['code'] = c
            elif v in DESC_HDRS and 'desc' not in roles: roles['desc'] = c
            elif v in KES_HDRS and 'kes' not in roles: roles['kes'] = c
            elif v in AMT_POST and 'amt_post' not in roles: roles['amt_post'] = c
            elif v in AMT_PLAIN and 'amt_plain' not in roles: roles['amt_plain'] = c
        if 'desc' in roles and ('amt_post' in roles or 'amt_plain' in roles):
            return r, roles
    return None, None

SKIP_DESC = {'assets','non-current assets','current assets','fixed assets','liabilities','equity',
    'equity and liabilities','owners equity','current liabilities','non-current liabilities','income',
    'expenses','total','total assets','total equity','total equity and liabilities','total income',
    'total expenses',''}

def read_entity(ws, descriptive):
    hr, roles = find_header(ws)
    if hr is None: return None
    ccode = roles.get('code'); cdesc = roles['desc']
    camt = roles.get('amt_post') or roles.get('amt_plain')   # local amount (post-accrual preferred)
    ckes = roles.get('kes')
    out = {}   # key -> [description, local_amount, kes_amount]
    for r in range(hr + 1, ws.max_row + 1):
        desc = ws.cell(r, cdesc).value
        nd = norm(desc)
        if nd in SKIP_DESC or nd.startswith('total') or nd.startswith('period'): continue
        loc = num(ws.cell(r, camt).value)
        kes = num(ws.cell(r, ckes).value) if ckes else None
        if loc is None and kes is None: continue
        code = None
        if ccode is not None and ws.cell(r, ccode).value not in (None, ''):
            code = str(ws.cell(r, ccode).value).strip()
        if not code:                       # descriptive / blank code -> synthetic, stable code from description
            if not desc: continue
            code = None                    # resolved per entity below (bridge or synthetic)
        key = code if code else ('~DESC~' + str(desc).strip())
        if key not in out:
            out[key] = [str(desc).strip() if desc else '', 0.0, 0.0 if ckes else None]
        out[key][1] += (loc or 0.0)
        if ckes is not None: out[key][2] += (kes or 0.0)
    return out

def load_existing(seedfile, code_field, amt_field):
    """cumulative amount per code, and the set of posting dates present."""
    cum = {}; dates = set(); header = None; desc = {}
    if os.path.exists(seedfile):
        with open(seedfile, newline='') as f:
            rd = csv.DictReader(f); header = rd.fieldnames
            for row in rd:
                c = row[code_field]; cum[c] = cum.get(c, 0.0) + float(row[amt_field] or 0)
                desc.setdefault(c, row.get('Description', ''))
                if row.get('Posting_Date'): dates.add(row['Posting_Date'])
    return cum, dates, header, desc

def bridge_descriptive(newrows, seed_desc):
    """For a descriptive tab whose seed is code-based: map '~DESC~x' keys to the
    existing seed code by description; unmatched -> synthetic stable code."""
    d2c = {}
    for code, d in seed_desc.items():
        d2c.setdefault(norm(d), code)
    fixed = {}
    for key, val in newrows.items():
        if key.startswith('~DESC~'):
            d = key[6:]
            code = d2c.get(norm(d)) or ('X-' + hashlib.md5(d.encode('utf-8')).hexdigest()[:10].upper())
            fixed[code] = val
        else:
            fixed[key] = val
    return fixed

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('workbook')
    ap.add_argument('--write', action='store_true')
    args = ap.parse_args()

    period, pend = parse_period(args.workbook)
    print(f'File: {os.path.basename(args.workbook)}')
    print(f'Period: {period}  (movements dated {pend})\n')
    wb = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)

    summary = []
    for tab, (seed, cur, descriptive) in ENTITIES.items():
        if tab not in wb.sheetnames: continue
        new = read_entity(wb[tab], descriptive)
        if not new:
            summary.append((seed, tab, 'no header found', 0, 0)); continue

        is_mena = (seed == 'mena')
        seedfile = os.path.join(BRONZE, f'gl_entry_{seed}.csv')
        code_field = 'Description' if is_mena else 'G_L_Account_No'
        loc_field  = 'Net_Debit_Credit' if is_mena else 'Amount'

        # prior cumulative per key (local + kes), and existing periods
        prior_loc, prior_kes, dates, header, seed_desc = {}, {}, set(), None, {}
        if os.path.exists(seedfile):
            with open(seedfile, newline='') as f:
                rd = csv.DictReader(f); header = rd.fieldnames
                for row in rd:
                    k = row[code_field]
                    inperiod = (row.get('Posting_Date') == pend)   # exclude target period (idempotent)
                    if not inperiod:
                        prior_loc[k] = prior_loc.get(k, 0.0) + float(row.get(loc_field) or 0)
                        if is_mena: prior_kes[k] = prior_kes.get(k, 0.0) + float(row.get('Amount_KES') or 0)
                    seed_desc.setdefault(k, row.get('Description', ''))
                    if row.get('Posting_Date'): dates.add(row['Posting_Date'])

        if not is_mena:
            new = bridge_descriptive(new, seed_desc)     # resolve descriptive/blank-code keys to seed codes

        # movements = new YTD - prior cumulative (covers new, changed, and disappeared accounts)
        keys = set(new) | set(prior_loc)
        movements = []
        for k in keys:
            nd, nloc, nkes = new.get(k, ['', 0.0, 0.0 if is_mena else None])
            mv  = round((nloc or 0.0) - prior_loc.get(k, 0.0), 2)
            mvk = round((nkes or 0.0) - prior_kes.get(k, 0.0), 2) if is_mena else None
            if abs(mv) < 0.005 and (not is_mena or abs(mvk) < 0.005): continue
            desc = nd or seed_desc.get(k, '')
            movements.append((k, desc, mv, mvk))

        summary.append((seed, tab, cur + (' [descriptive]' if descriptive else ''), len(new), len(movements)))

        if args.write:
            write_movements(seedfile, header, is_mena, pend, period, movements, dates)

    # FX (best effort)
    fx_note = update_fx(wb, period) if args.write else fx_preview(wb, period)

    print(f'{"entity":9}{"tab":12}{"currency":22}{"accounts":>10}{"movements":>11}')
    for seed, tab, cur, na, nm in summary:
        print(f'{seed:9}{tab:12}{cur:22}{na:>10}{nm:>11}')
    print('\n' + fx_note)
    if not args.write:
        print('\nDRY RUN — nothing written. Re-run with --write to update the seeds.')
    else:
        print('\nWROTE seeds. Next: dbt seed --full-refresh && dbt build')

def write_movements(seedfile, header, is_mena, pend, period, movements, dates):
    # drop existing rows for this period (idempotent)
    rows = []
    if os.path.exists(seedfile):
        with open(seedfile, newline='') as f:
            rd = csv.DictReader(f); header = rd.fieldnames
            rows = [r for r in rd if r.get('Posting_Date') != pend]
    if is_mena:
        header = header or ['Posting_Date','Description','Net_Debit_Credit','Amount_KES']
        with open(seedfile, 'w', newline='') as f:
            w = csv.DictWriter(f, fieldnames=header); w.writeheader(); w.writerows(rows)
            for code, desc, mv, kesmv in sorted(movements):
                w.writerow({'Posting_Date': pend, 'Description': desc,
                            'Net_Debit_Credit': mv, 'Amount_KES': round(kesmv or 0.0, 2)})
        return
    header = header or STD_HEADER
    maxno = max([int(r['Entry_No']) for r in rows], default=0)
    with open(seedfile, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=header); w.writeheader(); w.writerows(rows)
        for i, (code, desc, mv, _) in enumerate(sorted(movements), 1):
            no = maxno + i
            rec = {k: '' for k in header}
            rec.update({'Entry_No': no, 'Posting_Date': pend, 'Document_No': f'WB-{period}',
                        'Transaction_No': no, 'G_L_Account_No': code, 'Description': desc, 'Amount': mv,
                        'Debit_Amount': mv if mv > 0 else 0, 'Credit_Amount': -mv if mv < 0 else 0,
                        'Source_Code': 'WORKBOOK', 'Dimension_Set_ID': 0, 'User_ID': 'WORKBOOK_LOAD',
                        'Reversed': 'false'})
            w.writerow(rec)

RATE_CURR = {'rwanda franc':'RWF','malawi kwacha':'MWK','ugx shs':'UGX','nigeria naira':'NGN',
             'tanzania shs':'TZS','usd':'USD','aed':'AED'}
def _read_rates(wb, period):
    if 'Rates' not in wb.sheetnames: return None
    ws = wb['Rates']
    # header row with currency names; the period's closing row is the one whose month == period
    hdr = None; cols = {}
    for r in range(1, min(12, ws.max_row) + 1):
        vals = {c: norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        hit = {c: RATE_CURR[v] for c, v in vals.items() if v in RATE_CURR}
        if len(hit) >= 3: hdr = r; cols = hit; break
    if not hdr: return None
    closing = None
    for r in range(hdr + 1, ws.max_row + 1):
        d = ws.cell(r, 1).value
        if isinstance(d, datetime.datetime) and f'{d.year:04d}-{d.month:02d}' == period:
            closing = r
    if closing is None: return None
    out = {'KES': 1.0}
    for c, cur in cols.items():
        v = num(ws.cell(closing, c).value)
        if v: out[cur] = v
    return out

def fx_preview(wb, period):
    r = _read_rates(wb, period)
    return f'FX (Rates tab, {period}): {r}' if r else 'FX: no usable Rates tab row for this period (add fx_rate.csv manually).'

def update_fx(wb, period):
    r = _read_rates(wb, period)
    if not r: return 'FX: no usable Rates tab row — add fx_rate.csv rows for this period manually.'
    fxfile = os.path.join(REF, 'fx_rate.csv')
    rows = []
    if os.path.exists(fxfile):
        with open(fxfile, newline='') as f:
            rd = csv.DictReader(f); fld = rd.fieldnames
            rows = [x for x in rd if x['period'] != period]
    else:
        fld = ['currency','period','rate_type','rate_to_kes','rate_source']
    with open(fxfile, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fld); w.writeheader(); w.writerows(rows)
        for cur, val in r.items():
            src = 'CBK' if cur == 'KES' else 'RATES_TAB'
            for rt in ('CLOSING', 'AVERAGE'):
                w.writerow({'currency': cur, 'period': period, 'rate_type': rt,
                            'rate_to_kes': f'{val:.10f}', 'rate_source': src})
    return f'FX: wrote {len(r)} currencies for {period} to fx_rate.csv (verify rates).'

if __name__ == '__main__':
    main()
